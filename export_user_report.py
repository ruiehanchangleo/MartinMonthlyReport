#!/usr/bin/env python3
"""
Export user statistics debug report to Excel.
Shows who worked on what, including languages and excluded users.
"""

import json
import sys
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from generate_report import XTMReportGenerator

def main():
    # Load config
    config_path = 'xtm_config.json'
    with open(config_path, 'r') as f:
        config = json.load(f)

    # Create report generator
    generator = XTMReportGenerator(config_path, auto_send=False)

    print("=" * 80)
    print("XTM USER STATISTICS - EXCEL EXPORT")
    print("=" * 80)
    print(f"\nExcluded users: leo.chang@familysearch.org, LeoAdmin")
    print(f"Report period: {generator.report_month}")
    print("\n")

    # Get all projects
    projects = generator.get_projects()
    print(f"Total projects: {len(projects)}\n")
    print("Processing project data...")

    excluded_user_found = False
    projects_with_excluded = []
    total_users_by_email = {}
    all_project_details = []

    # Process each project
    for idx, project in enumerate(projects, 1):
        if idx % 20 == 0:
            print(f"  Processing project {idx}/{len(projects)}...")

        project_id = project['id']
        project_name = project['name']

        # Get statistics for this project (pass empty list to not exclude anyone)
        stats_list = generator.get_project_statistics(project_id, excluded_users=[])

        if not stats_list:
            continue

        project_has_excluded = False

        # Process each target language
        for lang_stats in stats_list:
            target_language = lang_stats.get('targetLanguage', 'Unknown')
            language_name = generator._locale_to_language_name(target_language)
            users_statistics = lang_stats.get('usersStatistics', [])

            for user_stats in users_statistics:
                username = user_stats.get('username', 'Unknown')

                # Track if this is an excluded user
                is_excluded = username.lower() in ['leo.chang@familysearch.org', 'leoadmin']
                if is_excluded:
                    excluded_user_found = True
                    project_has_excluded = True

                # Count users
                if username not in total_users_by_email:
                    total_users_by_email[username] = {
                        'projects': set(),
                        'total_words': 0,
                        'is_excluded': is_excluded,
                        'languages': {}
                    }

                total_users_by_email[username]['projects'].add(project_id)

                # Get workflow steps
                steps_statistics = user_stats.get('stepsStatistics', [])

                for step_stat in steps_statistics:
                    step_name = step_stat.get('workflowStepName', '')
                    clean_step_name = ''.join([c for c in step_name if not c.isdigit()]).strip()
                    jobs_statistics = step_stat.get('jobsStatistics', [])

                    step_total_words = 0
                    for job_stat in jobs_statistics:
                        source_stats = job_stat.get('sourceStatistics', {})
                        words = source_stats.get('totalWords', 0)
                        step_total_words += words

                    if step_total_words > 0:
                        # Track language totals for the user
                        if language_name not in total_users_by_email[username]['languages']:
                            total_users_by_email[username]['languages'][language_name] = 0
                        total_users_by_email[username]['languages'][language_name] += step_total_words
                        total_users_by_email[username]['total_words'] += step_total_words

                        # Store project detail
                        all_project_details.append({
                            'project_id': project_id,
                            'project_name': project_name,
                            'user': username,
                            'is_excluded': is_excluded,
                            'language': language_name,
                            'workflow_step': clean_step_name,
                            'words': step_total_words
                        })

        if project_has_excluded:
            projects_with_excluded.append(project_name)

    print(f"\n✓ Processed all {len(projects)} projects")
    print(f"✓ Found {len(total_users_by_email)} unique users")
    print(f"✓ Collected {len(all_project_details)} work records")

    # Create Excel workbook
    print("\nCreating Excel workbook...")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    excluded_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: User Summary
    ws_summary = wb.create_sheet("User Summary", 0)
    ws_summary.append(["User Email", "Projects", "Total Words", "Languages", "Top Language", "Status"])

    # Apply header formatting
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Add user data
    row_idx = 2
    for username, data in sorted(total_users_by_email.items(), key=lambda x: x[1]['total_words'], reverse=True):
        status = "⚠️ EXCLUDED" if data['is_excluded'] else "Active"

        # Get top language
        if data['languages']:
            top_lang = max(data['languages'].items(), key=lambda x: x[1])
            top_language = f"{top_lang[0]} ({top_lang[1]:,} words)"
        else:
            top_language = "N/A"

        ws_summary.append([
            username,
            len(data['projects']),
            data['total_words'],
            len(data['languages']),
            top_language,
            status
        ])

        # Highlight excluded users
        if data['is_excluded']:
            for col in range(1, 7):
                ws_summary.cell(row=row_idx, column=col).fill = excluded_fill

        row_idx += 1

    # Format columns
    ws_summary.column_dimensions['A'].width = 45
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 12
    ws_summary.column_dimensions['E'].width = 35
    ws_summary.column_dimensions['F'].width = 12

    # Sheet 2: Languages by User
    ws_languages = wb.create_sheet("Languages by User", 1)
    ws_languages.append(["User Email", "Language", "Words", "Status"])

    for cell in ws_languages[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row_idx = 2
    for username, data in sorted(total_users_by_email.items()):
        status = "⚠️ EXCLUDED" if data['is_excluded'] else "Active"

        if data['languages']:
            for language, words in sorted(data['languages'].items(), key=lambda x: x[1], reverse=True):
                ws_languages.append([username, language, words, status])

                if data['is_excluded']:
                    for col in range(1, 5):
                        ws_languages.cell(row=row_idx, column=col).fill = excluded_fill

                row_idx += 1

    ws_languages.column_dimensions['A'].width = 45
    ws_languages.column_dimensions['B'].width = 30
    ws_languages.column_dimensions['C'].width = 15
    ws_languages.column_dimensions['D'].width = 12

    # Sheet 3: Project Details
    ws_details = wb.create_sheet("Project Details", 2)
    ws_details.append(["Project ID", "Project Name", "User", "Language", "Workflow Step", "Words", "Status"])

    for cell in ws_details[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row_idx = 2
    for detail in sorted(all_project_details, key=lambda x: (x['project_id'], x['user'], x['language'])):
        status = "⚠️ EXCLUDED" if detail['is_excluded'] else "Active"

        ws_details.append([
            detail['project_id'],
            detail['project_name'],
            detail['user'],
            detail['language'],
            detail['workflow_step'],
            detail['words'],
            status
        ])

        if detail['is_excluded']:
            for col in range(1, 8):
                ws_details.cell(row=row_idx, column=col).fill = excluded_fill

        row_idx += 1

    ws_details.column_dimensions['A'].width = 12
    ws_details.column_dimensions['B'].width = 60
    ws_details.column_dimensions['C'].width = 45
    ws_details.column_dimensions['D'].width = 30
    ws_details.column_dimensions['E'].width = 20
    ws_details.column_dimensions['F'].width = 12
    ws_details.column_dimensions['G'].width = 12

    # Enable AutoFilter for all sheets
    ws_summary.auto_filter.ref = f"A1:F{ws_summary.max_row}"
    ws_languages.auto_filter.ref = f"A1:D{ws_languages.max_row}"
    ws_details.auto_filter.ref = f"A1:G{ws_details.max_row}"

    # Save the workbook
    output_dir = Path(config['onedrive_path'])
    timestamp = datetime.now().strftime('%Y%m%d')
    output_file = output_dir / f"XTM_User_Statistics_{timestamp}.xlsx"

    wb.save(output_file)

    print(f"\n{'=' * 80}")
    print("EXPORT COMPLETE")
    print(f"{'=' * 80}")
    print(f"\n✓ Excel file saved to: {output_file}")
    print(f"\n📊 Sheets created:")
    print(f"   1. User Summary - Overview of all users with totals")
    print(f"   2. Languages by User - Language breakdown per user")
    print(f"   3. Project Details - Detailed project assignments")
    print(f"\n📈 Statistics:")
    print(f"   - Total users: {len(total_users_by_email)}")
    print(f"   - Total projects: {len(projects)}")
    print(f"   - Total work records: {len(all_project_details)}")

    if excluded_user_found:
        print(f"\n⚠️  EXCLUDED USERS FOUND!")
        print(f"   Excluded users (leo.chang@familysearch.org or LeoAdmin) worked on {len(projects_with_excluded)} projects")
    else:
        print(f"\n✓ Excluded users (leo.chang@familysearch.org, LeoAdmin) NOT found in any projects")

    print(f"\n{'=' * 80}\n")

if __name__ == '__main__':
    main()
