#!/usr/bin/env python3
"""
XTM Monthly Report Generator

This script connects to XTM Cloud API, retrieves monthly metrics,
generates an Excel report, and prepares it for email distribution.
"""

import json
import logging
import sys
import subprocess
import platform
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Any, Callable
from functools import wraps
import requests



# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('xtm_report.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


def retry_with_backoff(max_attempts=5, initial_delay=1, backoff_factor=2, max_delay=60):
    """
    Decorator that retries a function with exponential backoff.

    Args:
        max_attempts: Maximum number of retry attempts
        initial_delay: Initial delay in seconds
        backoff_factor: Multiplier for delay after each attempt
        max_delay: Maximum delay between retries in seconds
    """
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            delay = initial_delay
            last_exception = None

            for attempt in range(1, max_attempts + 1):
                try:
                    return func(*args, **kwargs)
                except requests.exceptions.Timeout as e:
                    last_exception = e
                    logger.warning(f"Timeout on attempt {attempt}/{max_attempts} for {func.__name__}: {e}")
                except requests.exceptions.ConnectionError as e:
                    last_exception = e
                    logger.warning(f"Connection error on attempt {attempt}/{max_attempts} for {func.__name__}: {e}")
                except requests.exceptions.HTTPError as e:
                    last_exception = e
                    # Don't retry on 4xx errors (client errors) except 429 (rate limit)
                    if hasattr(e, 'response') and e.response is not None:
                        status_code = e.response.status_code
                        if 400 <= status_code < 500 and status_code != 429:
                            logger.error(f"Client error {status_code} for {func.__name__}, not retrying: {e}")
                            raise
                    logger.warning(f"HTTP error on attempt {attempt}/{max_attempts} for {func.__name__}: {e}")
                except requests.exceptions.RequestException as e:
                    last_exception = e
                    logger.warning(f"Request error on attempt {attempt}/{max_attempts} for {func.__name__}: {e}")
                except Exception as e:
                    # Don't retry on unexpected exceptions
                    logger.error(f"Unexpected error in {func.__name__}: {e}")
                    raise

                # If this wasn't the last attempt, wait before retrying
                if attempt < max_attempts:
                    wait_time = min(delay, max_delay)
                    logger.info(f"Retrying {func.__name__} in {wait_time} seconds...")
                    time.sleep(wait_time)
                    delay *= backoff_factor

            # All attempts failed
            logger.error(f"All {max_attempts} attempts failed for {func.__name__}")
            raise last_exception

        return wrapper
    return decorator


class XTMReportGenerator:
    """Main class for generating XTM monthly reports."""

    # Locale code to language name mapping
    LOCALE_TO_LANGUAGE = {
        'en_US': 'English (USA)',
        'ar_AA': 'Arabic',
        'ar_AE': 'Arabic (UAE)',
        'ar_EG': 'Arabic (Egypt)',
        'ar_SA': 'Arabic (Saudi Arabia)',
        'bg_BG': 'Bulgarian',
        'ceb': 'Cebuano',
        'cs_CZ': 'Czech',
        'da_DK': 'Danish',
        'de_DE': 'German',
        'el_CY': 'Greek (Cyprus)',
        'el_GR': 'Greek',
        'en_US': 'English (US)',
        'en_GB': 'English (UK)',
        'es_ES': 'Spanish',
        'es_MX': 'Spanish (Mexico)',
        'et_EE': 'Estonian',
        'fa_IR': 'Persian',
        'fi_FI': 'Finnish',
        'fj_FJ': 'Fijian',
        'fr_FR': 'French (France)',
        'hr_BA': 'Croatian (Bosnia)',
        'hr_HR': 'Croatian',
        'ht_HT': 'Haitian Creole',
        'hu_HU': 'Hungarian',
        'hy_AM': 'Armenian',
        'id_ID': 'Indonesian',
        'is_IS': 'Icelandic',
        'it_IT': 'Italian',
        'ja_JP': 'Japanese',
        'ka_GE': 'Georgian',
        'kk_KZ': 'Kazakh',
        'km_KH': 'Khmer',
        'ko_KR': 'Korean',
        'lo_LA': 'Lao',
        'lt_LT': 'Lithuanian',
        'lv_LV': 'Latvian',
        'mg_MG': 'Malagasy',
        'mk_MK': 'Macedonian',
        'mn_MN': 'Mongolian',
        'ms_MY': 'Malay',
        'nl_NL': 'Dutch',
        'no_NO': 'Norwegian',
        'nb_NO': 'Norwegian',
        'pl_PL': 'Polish',
        'pt_BR': 'Portuguese (Brazil)',
        'pt_PT': 'Portuguese (Portugal)',
        'ro_RO': 'Romanian',
        'ru_RU': 'Russian',
        'sk_SK': 'Slovak',
        'sl_SI': 'Slovenian',
        'sm_WS': 'Samoan',
        'sq_AL': 'Albanian',
        'sr_RS': 'Serbian',
        'sv_SE': 'Swedish',
        'sw_KE': 'Swahili (Kenya)',
        'sw_TZ': 'Swahili (Tanzania)',
        'swa': 'Swahili',
        'th_TH': 'Thai',
        'tl_PH': 'Tagalog',
        'tl': 'Tagalog',
        'fil_PH': 'Tagalog',
        'fil': 'Tagalog',
        'to_TO': 'Tongan',
        'tr_TR': 'Turkish',
        'ty': 'Tahitian',
        'uk_UA': 'Ukrainian',
        'ur_IN': 'Urdu',
        'vi_VN': 'Vietnamese',
        'goyu': 'Chinese',
        'zh_CN': 'Chinese (Simplified)',
        'zh_HK': 'Chinese (Hong Kong)',
        'zh_TW': 'Chinese (Traditional)',
        'es_419': 'Spanish (Latin America)',
        'tr': 'Turkish',
    }

    # Accounts to exclude, matched against the XTM `username` field ONLY (case-insensitive).
    # Display names, firstName/lastName, and email domains are intentionally NOT considered.
    EXCLUDED_USERS = [
        "BreaB@familysearch.org",
        "BSP",
        "BSP_Tester",
        "LeoAdmin",
        "lingoport",
        "MartinADMIN",
        "FS Linguist",
        "thelinguist",
        "testuser@gmail.com",
        "DummyUser",
        "xtmsupport",
        "XTM TM import API",
    ]

    def __init__(self, config_path: str = "xtm_config.json", auto_send: bool = False, weekly: bool = False):
        """Initialize the report generator with configuration."""
        self.config = self._load_config(config_path)
        self.auto_send = auto_send
        self.weekly = weekly
        self.base_url = self.config['base_url']
        self.headers = {
            'Authorization': f"{self.config['auth_type']} {self.config['auth_token']}",
            'Content-Type': 'application/json'
        }
        self.report_date = datetime.now()
        self._volunteers_cache = None  # memoized roster (see get_volunteers)
        self._volunteer_hours = None   # per-volunteer active hours, current period (see _compute_volunteer_hours)
        self._volunteer_hours_ytd = None  # per-volunteer active hours, year-to-date totals (monthly reports only)
        self._volunteer_hours_ytd_breakdown = None  # YTD active hours per volunteer per month (for the trend chart)

        if self.weekly:
            # Calculate previous 7 days for weekly report
            # Week ends on Sunday (yesterday if today is Monday, or last Sunday)
            end_date = self.report_date - timedelta(days=1)  # Yesterday
            start_date = end_date - timedelta(days=6)  # 7 days total including end_date

            self.report_start_date = start_date
            self.report_end_date = end_date
            self.report_period = f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
            self.report_week_label = f"Week of {start_date.strftime('%Y-%m-%d')}"

            # For weekly reports, we don't use month-based ranges
            self.report_month = None
            self.report_month_name = None
            self.ytd_start_month = None
            self.ytd_end_month = None
        else:
            # Calculate previous month for the monthly report
            first_day_current_month = self.report_date.replace(day=1)
            last_day_previous_month = first_day_current_month - timedelta(days=1)

            self.report_month = last_day_previous_month.strftime('%Y-%m')
            self.report_month_name = last_day_previous_month.strftime('%B %Y')

            # Get year-to-date range (January 1 to end of previous month)
            self.ytd_start_month = last_day_previous_month.strftime('%Y-01')
            self.ytd_end_month = self.report_month

            # Monthly reports don't use these
            self.report_start_date = None
            self.report_end_date = None
            self.report_period = None
            self.report_week_label = None

    def _load_config(self, config_path: str) -> Dict:
        """Load configuration from JSON file."""
        try:
            config_file = Path(config_path)
            if not config_file.exists():
                raise FileNotFoundError(f"Configuration file not found: {config_path}")

            with open(config_path, 'r') as f:
                config = json.load(f)

            # Validate required configuration keys
            required_keys = ['base_url', 'auth_type', 'auth_token', 'onedrive_path', 'email_recipients']
            missing_keys = [key for key in required_keys if key not in config]
            if missing_keys:
                raise ValueError(f"Missing required config keys: {', '.join(missing_keys)}")

            # Validate auth token is not empty
            if not config.get('auth_token') or config['auth_token'].strip() == '':
                raise ValueError("Authentication token is empty")

            # Validate email recipients
            if not isinstance(config.get('email_recipients'), list) or not config['email_recipients']:
                logger.warning("No email recipients configured")

            logger.info("Configuration loaded and validated successfully")
            return config
        except Exception as e:
            logger.error(f"Failed to load config: {e}")
            raise

    def _locale_to_language_name(self, locale_code: str) -> str:
        """Convert locale code to base language name, combining variants."""
        name = self.LOCALE_TO_LANGUAGE.get(locale_code, locale_code)
        # Strip parenthetical variants: "Portuguese (Brazil)" -> "Portuguese"
        if '(' in name:
            name = name.split('(')[0].strip()
        return name

    def _run_health_checks(self) -> bool:
        """Run comprehensive health checks before starting report generation."""
        logger.info("Running health checks...")
        all_passed = True

        # 1. Check API connectivity
        try:
            logger.info("Checking API connectivity...")
            test_response = self._make_request('projects', params={'count': 1})
            logger.info("✓ API connectivity check passed")
        except Exception as e:
            logger.error(f"✗ API connectivity check failed: {e}")
            all_passed = False

        # 2. Check output directory
        try:
            logger.info("Checking output directory...")
            output_dir = Path(self.config['onedrive_path'])

            # Check if directory exists or can be created
            if not output_dir.exists():
                try:
                    output_dir.mkdir(parents=True, exist_ok=True)
                    logger.info(f"✓ Created output directory: {output_dir}")
                except Exception as e:
                    logger.warning(f"⚠ Cannot create primary output directory, will use fallback: {e}")

            # Check write permissions
            if output_dir.exists():
                test_file = output_dir / '.write_test'
                try:
                    test_file.write_text('test')
                    test_file.unlink()
                    logger.info(f"✓ Output directory is writable: {output_dir}")
                except Exception as e:
                    logger.warning(f"⚠ Output directory not writable, will use fallback: {e}")
            else:
                logger.warning(f"⚠ Output directory does not exist, will use fallback locations")

        except Exception as e:
            logger.warning(f"⚠ Output directory check failed, will use fallback: {e}")

        # 3. Check disk space (warn if less than 100MB free)
        try:
            logger.info("Checking disk space...")
            import shutil
            stat = shutil.disk_usage(Path.home())
            free_mb = stat.free / (1024 * 1024)
            if free_mb < 100:
                logger.warning(f"⚠ Low disk space: {free_mb:.0f}MB free")
            else:
                logger.info(f"✓ Sufficient disk space: {free_mb:.0f}MB free")
        except Exception as e:
            logger.warning(f"⚠ Could not check disk space: {e}")

        # 4. Check required Python packages
        try:
            logger.info("Checking required packages...")
            import openpyxl
            import requests
            logger.info("✓ All required packages are available")
        except ImportError as e:
            logger.error(f"✗ Missing required package: {e}")
            all_passed = False

        # 5. Validate date calculations
        try:
            logger.info("Validating date calculations...")
            if self.weekly:
                if not self.report_start_date or not self.report_end_date:
                    raise ValueError("Invalid weekly date calculations")
                logger.info(f"✓ Weekly report period: {self.report_period}")
            else:
                if not self.report_month or not self.ytd_start_month:
                    raise ValueError("Invalid date calculations")
                logger.info(f"✓ Report period: {self.report_month_name}")
                logger.info(f"✓ YTD period: {self.ytd_start_month} to {self.ytd_end_month}")
        except Exception as e:
            logger.error(f"✗ Date validation failed: {e}")
            all_passed = False

        if all_passed:
            logger.info("✓ All critical health checks passed")
        else:
            logger.warning("⚠ Some health checks failed, but will attempt to continue")

        return all_passed

    @retry_with_backoff(max_attempts=5, initial_delay=2, backoff_factor=2, max_delay=60)
    def _make_request(self, endpoint: str, method: str = 'GET', params: Dict = None, data: Dict = None) -> Any:
        """Make API request to XTM with error handling."""
        url = f"{self.base_url}/{endpoint}"
        try:
            logger.info(f"Making {method} request to {endpoint}")
            if method == 'GET':
                response = requests.get(url, headers=self.headers, params=params, timeout=60)
            elif method == 'POST':
                response = requests.post(url, headers=self.headers, json=data, timeout=60)
            else:
                raise ValueError(f"Unsupported method: {method}")

            # Log the xtm-trace-id header for support purposes
            trace_id = response.headers.get('xtm-trace-id', 'N/A')
            logger.info(f"XTM Trace ID: {trace_id} | Endpoint: {endpoint} | Status: {response.status_code}")

            response.raise_for_status()
            return response.json() if response.content else {}
        except requests.exceptions.RequestException as e:
            # Also log trace ID on error if available
            trace_id = 'N/A'
            if hasattr(e, 'response') and e.response is not None:
                trace_id = e.response.headers.get('xtm-trace-id', 'N/A')
                logger.error(f"API request failed | XTM Trace ID: {trace_id} | Error: {e}")
            else:
                logger.error(f"API request failed: {e}")
            raise

    # XTM list endpoints cap each response at pageSize (max 1000) and require
    # paging to get everything. _fetch_all_pages loops until a short page so we
    # never silently truncate (the bug that hid 800+ projects). MAX_PAGES is a
    # safety stop against an unexpected infinite loop.
    MAX_PAGES = 100

    def _fetch_all_pages(self, endpoint: str, params: Dict = None, page_size: int = 1000) -> List[Dict]:
        """Fetch every page of a paginated XTM list endpoint into one list."""
        results = []
        params = dict(params or {})
        for page in range(1, self.MAX_PAGES + 1):
            params['page'] = page
            params['pageSize'] = page_size
            batch = self._make_request(endpoint, params=params)
            if not isinstance(batch, list) or not batch:
                break
            results.extend(batch)
            if len(batch) < page_size:
                break
            if page == self.MAX_PAGES:
                logger.warning(f"{endpoint}: hit MAX_PAGES={self.MAX_PAGES} cap; "
                               f"results may be truncated — raise MAX_PAGES")
        return results

    def get_projects(self, status: str = None, modified_from: str = None) -> List[Dict]:
        """Retrieve projects from XTM, paging through ALL results (no 1000 cap).

        modified_from (ISO 'YYYY-MM-DD'): only fetch projects modified on/after
        this date. Safe for reporting because any project with work completed in
        a period was necessarily modified at/after that completion, so its
        modificationDate >= period start — filtering by the period start never
        drops in-period work, but skips the many archived projects untouched
        since then (big speedup). Leave None to fetch everything (e.g. snapshots)."""
        try:
            params = {}
            if status:
                params['status'] = status
            if modified_from:
                params['modifiedDateFrom'] = modified_from
            projects = self._fetch_all_pages('projects', params=params)
            logger.info(f"Retrieved {len(projects)} projects"
                        + (f" modified since {modified_from}" if modified_from else ""))
            return projects
        except Exception as e:
            logger.error(f"Failed to get projects: {e}")
            return []

    def get_project_metrics(self, project_id: int):
        """Get detailed metrics for a specific project (returns list)."""
        try:
            metrics = self._make_request(f'projects/{project_id}/metrics')
            # Metrics is a list of target languages
            if isinstance(metrics, list):
                return metrics
            elif isinstance(metrics, dict):
                return [metrics]  # Wrap single dict in list for consistency
            return []
        except Exception as e:
            logger.warning(f"Failed to get metrics for project {project_id}: {e}")
            return []

    def get_project_metrics_data(self, project_id: int):
        """
        Get project metrics showing wordsDone for each workflow step.
        Returns metrics for all languages in the project with wordsDone by workflow step.
        This shows CURRENT progress, not historical monthly completions.
        """
        try:
            metrics = self._make_request(f'projects/{project_id}/metrics')
            if not isinstance(metrics, list):
                return []
            return metrics
        except Exception as e:
            logger.warning(f"Failed to get metrics for project {project_id}: {e}")
            return []

    def get_project_status_with_steps(self, project_id: int):
        """
        Get project status with workflow step completion dates.
        This endpoint preserves finishDate even when work is reopened.
        Returns job-level data with step-level finish dates.
        """
        try:
            status = self._make_request(f'projects/{project_id}/status?fetchLevel=STEPS')
            if not isinstance(status, dict):
                return {}
            return status
        except Exception as e:
            logger.warning(f"Failed to get status for project {project_id}: {e}")
            return {}

    def _is_excluded_user(self, user_stat: Dict) -> bool:
        """Check if a user should be excluded, matching the `username` field only (case-insensitive)."""
        username = user_stat.get('username', '').lower()
        if not username:
            return False

        excluded_lower = [u.lower() for u in self.EXCLUDED_USERS]
        return username in excluded_lower

    @staticmethod
    def _resolve_user_name(user_stat: Dict) -> str:
        """
        Extract the best display name from a user statistics entry.
        Uses userDisplayName as-is (no reversal) since XTM API doesn't provide firstName/lastName.
        """
        # Use userDisplayName directly - DO NOT reverse the name
        # XTM displays names in various formats, we preserve them as-is
        username = user_stat.get('userDisplayName', user_stat.get('username', 'Unknown'))

        # Strip "generic " prefix from XTM display names
        if username.lower().startswith('generic '):
            username = username[8:]

        return username

    def _filter_excluded_from_stats(self, stats: List[Dict]) -> List[Dict]:
        """Remove excluded users from a raw /statistics list (drops languages left empty)."""
        if not isinstance(stats, list):
            return []

        filtered_stats = []
        for lang_stats in stats:
            users_statistics = lang_stats.get('usersStatistics', [])
            filtered_users = [
                user for user in users_statistics
                if not self._is_excluded_user(user)
            ]

            if filtered_users:
                lang_stats_copy = lang_stats.copy()
                lang_stats_copy['usersStatistics'] = filtered_users
                filtered_stats.append(lang_stats_copy)

        return filtered_stats

    def get_volunteers(self, force_refresh: bool = False) -> Dict[str, Dict]:
        """Return the authoritative volunteer roster from XTM: every user in
        /users that is NOT in EXCLUDED_USERS, keyed by lowercased username.

        Each entry is {'user': display name, 'username': login, 'languages':
        [target language names]} where languages come from the user's assigned
        /users/{id}/language-combinations. Used to list every volunteer in the
        user report — including those with no work this period (shown as zeros).

        The roster is memoized in-process and cached to .cache/volunteers.json
        for 24h so repeated runs don't re-issue one API call per user."""
        import time
        import re
        from concurrent.futures import ThreadPoolExecutor

        if self._volunteers_cache is not None:
            return self._volunteers_cache

        cache_path = Path(__file__).parent / ".cache" / "volunteers.json"
        cache_path.parent.mkdir(exist_ok=True)
        if not force_refresh and cache_path.exists():
            try:
                if (time.time() - cache_path.stat().st_mtime) < 24 * 3600:
                    with open(cache_path) as f:
                        self._volunteers_cache = json.load(f)
                    logger.info(f"Loaded {len(self._volunteers_cache)} volunteers from cache")
                    return self._volunteers_cache
            except Exception as e:
                logger.warning(f"Failed to read volunteer cache: {e}")

        try:
            users = self._fetch_all_pages('users')
        except Exception as e:
            logger.warning(f"Failed to fetch volunteer roster: {e}")
            return {}

        excluded_lower = {u.lower() for u in self.EXCLUDED_USERS}
        roster = [u for u in users
                  if u.get('id') and u.get('username', '').lower() not in excluded_lower]
        logger.info(f"Fetching language combinations for {len(roster)} volunteers...")

        def fetch_langs(u):
            try:
                resp = self._make_request(f"users/{u['id']}/language-combinations") or {}
                combos = resp.get('languageCombinations', [])
            except Exception:
                combos = []
            langs = sorted({self._locale_to_language_name(c.get('targetLanguage'))
                            for c in combos if c.get('targetLanguage')})
            return u, langs

        volunteers = {}
        with ThreadPoolExecutor(max_workers=10) as executor:
            for u, langs in executor.map(fetch_langs, roster):
                uname = u.get('username', '')
                # Drop XTM's "generic" token from names (mirrors _resolve_user_name).
                name = re.sub(r'\bgeneric\b', '', f"{u.get('firstName', '')} {u.get('lastName', '')}", flags=re.I)
                name = ' '.join(name.split()) or uname
                volunteers[uname.lower()] = {
                    'user': name,
                    'username': uname,
                    'languages': langs,
                }

        try:
            with open(cache_path, 'w') as f:
                json.dump(volunteers, f)
        except Exception as e:
            logger.warning(f"Failed to write volunteer cache: {e}")

        self._volunteers_cache = volunteers
        logger.info(f"Volunteer roster ready: {len(volunteers)} volunteers")
        return volunteers

    def _inject_zero_volunteers(self, user_dict: Dict, value_field: str):
        """Add zero-activity rows to a user statistics dict for every volunteer
        in the roster who did no work in the period. One row per assigned target
        language (from the roster); volunteers with no assigned languages get a
        single language-less row. Volunteers who did any work are left untouched
        (matched by username), so their real per-language rows are unaffected.

        `value_field` is the per-user data bucket to initialize empty:
        'workflow_steps' for current-period tables, 'months' for YTD."""
        roster = self.get_volunteers()
        if not roster:
            return
        present = {(ud.get('username') or '').lower()
                   for ud in user_dict.values() if ud.get('username')}
        for uname, info in roster.items():
            if uname in present:
                continue
            for lang in (info['languages'] or ['']):
                key = f"{info['user']}|{lang}"
                if key in user_dict:
                    continue
                user_dict[key] = {
                    'user': info['user'],
                    'username': info['username'],
                    'language': lang,
                    value_field: {},
                }

    def get_project_statistics_raw(self, project_id: int) -> List[Dict]:
        """Get raw per-user statistics for a project (no exclusion filtering applied)."""
        try:
            stats = self._make_request(f'projects/{project_id}/statistics')
            return stats if isinstance(stats, list) else []
        except Exception as e:
            logger.warning(f"Failed to get statistics for project {project_id}: {e}")
            return []

    def get_project_statistics(self, project_id: int):
        """Get detailed per-user statistics for a project, excluding specified users."""
        return self._filter_excluded_from_stats(self.get_project_statistics_raw(project_id))

    def get_workflow_steps(self, project_id: int) -> List[Dict]:
        """Get workflow steps for a project."""
        try:
            # Try different endpoints for workflow data
            try:
                steps = self._make_request(f'projects/{project_id}/jobs')
                return steps if isinstance(steps, list) else []
            except:
                # Fallback to user workflow steps
                steps = self._make_request(f'projects/{project_id}/workflow')
                return steps if isinstance(steps, list) else []
        except Exception as e:
            logger.debug(f"No workflow data available for project {project_id}: {e}")
            return []

    def get_users_workflow_steps(self) -> List[Dict]:
        """Get user workflow step assignments."""
        try:
            steps = self._make_request('users/workflow-steps')
            return steps if isinstance(steps, list) else []
        except Exception as e:
            logger.warning(f"Failed to get user workflow steps: {e}")
            return []

    def aggregate_monthly_data(self, start_month: str = None, end_month: str = None) -> Dict:
        """Aggregate data for the specified date range using completion dates."""
        if start_month is None:
            start_month = self.report_month
        if end_month is None:
            end_month = self.report_month

        logger.info(f"Aggregating data from {start_month} to {end_month}")

        # Parse date range
        from datetime import datetime
        start_date = datetime.strptime(start_month + '-01', '%Y-%m-%d')
        # End date is last day of end_month
        if end_month != start_month:
            end_year, end_month_num = map(int, end_month.split('-'))
            if end_month_num == 12:
                end_date = datetime(end_year + 1, 1, 1)
            else:
                end_date = datetime(end_year, end_month_num + 1, 1)
        else:
            year, month = map(int, start_month.split('-'))
            if month == 12:
                end_date = datetime(year + 1, 1, 1)
            else:
                end_date = datetime(year, month + 1, 1)

        logger.info(f"Date range: {start_date} to {end_date}")

        # Initialize data structures
        data = {
            'project_stats': {
                'total': 0,
                'completed': 0,
                'in_progress': 0,
                'pending': 0
            },
            'workflow_by_language': {},  # workflow metrics per language
            'user_statistics': {},  # User-level statistics
            'projects': []
        }

        # Only fetch projects modified since the period start — in-period work
        # always bumps modificationDate, so this can't drop real work but skips
        # archived projects untouched since then.
        projects = self.get_projects(modified_from=start_date.strftime('%Y-%m-%d'))

        # Fetch statistics and status for all projects in parallel
        from concurrent.futures import ThreadPoolExecutor, as_completed

        def fetch_project_data(project_id):
            try:
                stats = self.get_project_statistics(project_id)
                status = self.get_project_status_with_steps(project_id)
                return project_id, stats, status
            except Exception as e:
                logger.warning(f"Failed to get data for project {project_id}: {e}")
                return project_id, [], {}

        project_ids = [p.get('id') for p in projects if p.get('id')]
        project_map = {p.get('id'): p for p in projects if p.get('id')}
        project_results = {}

        logger.info(f"Fetching data for {len(project_ids)} projects using parallel requests...")
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = {executor.submit(fetch_project_data, pid): pid for pid in project_ids}
            for future in as_completed(futures):
                pid, stats_list, project_status = future.result()
                project_results[pid] = (stats_list, project_status)

        for project_id in project_ids:
            project = project_map[project_id]
            stats_list, project_status = project_results[project_id]

            # If the live statistics are empty (project archived), restore the
            # per-user breakdown from a snapshot taken while it was still active.
            # This preserves real user names instead of falling back to the
            # metrics-only "Archived User" path below.
            if not stats_list:
                snap_stats, snap_status = self._restore_stats_from_snapshot(project_id)
                if snap_stats:
                    stats_list = snap_stats
                    if not project_status:
                        project_status = snap_status
                    logger.info(f"Project {project_id} archived; restored per-user "
                                f"statistics from snapshot")

            # Build a map of completion dates from /status: {jobId: {stepName: finishDate}}
            # These finishDate values are preserved even when work is reopened
            # Use lowercase step names as keys to handle case mismatches between endpoints
            job_step_dates = {}
            if project_status:
                for job in project_status.get('jobs', []):
                    job_id = job.get('jobId')
                    if job_id:
                        job_step_dates[job_id] = {}
                        for step in job.get('steps', []):
                            step_name = step.get('workflowStepName', '')
                            finish_date = step.get('finishDate')
                            if finish_date and step_name:
                                # Store with lowercase key for case-insensitive lookup
                                job_step_dates[job_id][step_name.lower()] = finish_date

            # Process each target language
            # Check if statistics are empty (archived project)
            if not stats_list or len(stats_list) == 0:
                # Project is likely archived - try metrics endpoint as fallback
                logger.info(f"Project {project_id} has no statistics (likely archived), using metrics fallback")

                # Use metrics endpoint to get word counts
                metrics = self.get_project_metrics(project_id)

                if metrics and len(metrics) > 0:
                    project_total_words = 0
                    project_has_work_in_period = False
                    target_languages = []

                    for metric in metrics:
                        target_lang_code = metric.get('targetLanguage', 'unknown')
                        target_lang_name = self._locale_to_language_name(target_lang_code)

                        # Get step-level metrics
                        metrics_progress = metric.get('metricsProgress', {})

                        for step_name, step_metrics in metrics_progress.items():
                            words_done = step_metrics.get('wordsDone', 0)

                            if words_done > 0:
                                # Clean step name - remove digits and normalize to lowercase
                                clean_step_name = ''.join([c for c in step_name if not c.isdigit()]).strip().lower()

                                # Archived projects only expose a single cumulative
                                # wordsDone per language+step (no per-job words). To avoid
                                # double-counting, attribute that cumulative total to the
                                # ONE month in which the step was *fully* completed for this
                                # language — i.e. the latest finish date across the
                                # language's jobs for this step. The previous "any job that
                                # finished in this period" test re-added the same cumulative
                                # words in every month a job happened to finish, inflating
                                # YTD totals.
                                step_completed_in_period = False
                                step_finish_dates = []
                                if project_status:
                                    for job in project_status.get('jobs', []):
                                        if job.get('targetLanguage') == target_lang_code:
                                            for step in job.get('steps', []):
                                                if step.get('workflowStepName', '').lower() == step_name.lower():
                                                    finish_date_ts = step.get('finishDate')
                                                    if finish_date_ts:
                                                        step_finish_dates.append(finish_date_ts)

                                if step_finish_dates:
                                    completion_date = datetime.fromtimestamp(max(step_finish_dates) / 1000)
                                    if start_date <= completion_date < end_date:
                                        step_completed_in_period = True

                                if step_completed_in_period:
                                    project_has_work_in_period = True

                                    if target_lang_name not in target_languages:
                                        target_languages.append(target_lang_name)

                                    # Use generic "Archived User" for attribution,
                                    # qualified by language so each archived language
                                    # is distinguishable in the report
                                    username = f"Archived User ({target_lang_name})"

                                    # Track user statistics
                                    user_key = f"{username}|{target_lang_name}"
                                    if user_key not in data['user_statistics']:
                                        data['user_statistics'][user_key] = {
                                            'user': username,
                                            'language': target_lang_name,
                                            'workflow_steps': {}
                                        }

                                    # Add words to this user's workflow step
                                    if clean_step_name not in data['user_statistics'][user_key]['workflow_steps']:
                                        data['user_statistics'][user_key]['workflow_steps'][clean_step_name] = 0
                                    data['user_statistics'][user_key]['workflow_steps'][clean_step_name] += words_done

                                    # Create unique key for workflow step + language
                                    workflow_key = f"{clean_step_name} - {target_lang_name}"

                                    if workflow_key not in data['workflow_by_language']:
                                        data['workflow_by_language'][workflow_key] = {
                                            'workflow_step': clean_step_name,
                                            'language': target_lang_name,
                                            'words_done': 0,
                                            'words_to_be_done': 0,
                                            'projects': 0
                                        }

                                    data['workflow_by_language'][workflow_key]['words_done'] += words_done
                                    project_total_words += words_done

                    # Add project to stats if it has work in the target period
                    if project_has_work_in_period and project_total_words > 0:
                        data['project_stats']['total'] += 1
                        status = project.get('status', 'UNKNOWN')

                        if status == 'FINISHED':
                            data['project_stats']['completed'] += 1
                        elif status in ['IN_PROGRESS', 'STARTED']:
                            data['project_stats']['in_progress'] += 1
                        else:
                            data['project_stats']['pending'] += 1

                        # Mark projects for each language/workflow combination
                        for target_lang_name in target_languages:
                            for workflow_key in data['workflow_by_language'].keys():
                                if workflow_key.endswith(f" - {target_lang_name}"):
                                    data['workflow_by_language'][workflow_key]['projects'] = data['workflow_by_language'][workflow_key].get('projects', 0) + 1

                        # Store project summary
                        data['projects'].append({
                            'id': project_id,
                            'name': project.get('name', 'Unknown'),
                            'status': status,
                            'source_lang': 'en_US',
                            'target_langs': ', '.join(target_languages),
                            'total_words': project_total_words,
                            'created_date': None
                        })

                        logger.info(f"Added archived project {project_id} with {project_total_words} words from metrics")

                # Skip normal processing since we used metrics fallback
                continue

            if isinstance(stats_list, list) and stats_list:
                project_total_words = 0
                project_has_work_in_period = False
                target_languages = []
                source_lang = 'en_US'  # Default

                for lang_stats in stats_list:
                    target_lang_code = lang_stats.get('targetLanguage', 'unknown')
                    target_lang_name = self._locale_to_language_name(target_lang_code)

                    users_statistics = lang_stats.get('usersStatistics', [])

                    # Process each user's work (already filtered by get_project_statistics)
                    for user_stat in users_statistics:
                        steps_statistics = user_stat.get('stepsStatistics', [])

                        # Process each workflow step
                        for step_stat in steps_statistics:
                            step_name = step_stat.get('workflowStepName', '')
                            # Remove numbers from step name and normalize to lowercase
                            clean_step_name = ''.join([c for c in step_name if not c.isdigit()]).strip().lower()

                            jobs_statistics = step_stat.get('jobsStatistics', [])

                            # Process each job - FILTER BY STEP COMPLETION DATE
                            for job_stat in jobs_statistics:
                                include_this_job = False

                                # Get job ID to look up step finish date from /status
                                job_id = job_stat.get('jobId')

                                # Get the finish date for this specific step from /status endpoint
                                # This date is preserved even when work is reopened
                                # Use lowercase for case-insensitive lookup
                                finish_date_ts = None
                                if job_id and job_id in job_step_dates:
                                    finish_date_ts = job_step_dates[job_id].get(step_name.lower())

                                # Check if this step was completed in the target date range
                                if finish_date_ts:
                                    completion_date = datetime.fromtimestamp(finish_date_ts / 1000)
                                    if start_date <= completion_date < end_date:
                                        include_this_job = True
                                        logger.debug(f"Including job {job_id} step {step_name}: completed {completion_date}")

                                if include_this_job:
                                    # This job should be included!
                                    source_stats = job_stat.get('sourceStatistics', {})
                                    total_words = source_stats.get('totalWords', 0)

                                    if total_words > 0:
                                        project_has_work_in_period = True

                                        if target_lang_name not in target_languages:
                                            target_languages.append(target_lang_name)

                                        username = self._resolve_user_name(user_stat)

                                        # Track user statistics (user + language + workflow step)
                                        user_key = f"{username}|{target_lang_name}"
                                        if user_key not in data['user_statistics']:
                                            data['user_statistics'][user_key] = {
                                                'user': username,
                                                'username': user_stat.get('username', ''),
                                                'language': target_lang_name,
                                                'workflow_steps': {}
                                            }

                                        # Add words to this user's workflow step
                                        if clean_step_name not in data['user_statistics'][user_key]['workflow_steps']:
                                            data['user_statistics'][user_key]['workflow_steps'][clean_step_name] = 0
                                        data['user_statistics'][user_key]['workflow_steps'][clean_step_name] += total_words

                                        # Create unique key for workflow step + language
                                        workflow_key = f"{clean_step_name} - {target_lang_name}"

                                        if workflow_key not in data['workflow_by_language']:
                                            data['workflow_by_language'][workflow_key] = {
                                                'workflow_step': clean_step_name,
                                                'language': target_lang_name,
                                                'words_done': 0,
                                                'words_to_be_done': 0,
                                                'projects': 0
                                            }

                                        data['workflow_by_language'][workflow_key]['words_done'] += total_words
                                        project_total_words += total_words

                # Only add project to stats if it has work in the target period
                if project_has_work_in_period and project_total_words > 0:
                    # Update project stats
                    data['project_stats']['total'] += 1
                    status = project.get('status', 'UNKNOWN')

                    if status == 'FINISHED':
                        data['project_stats']['completed'] += 1
                    elif status in ['IN_PROGRESS', 'STARTED']:
                        data['project_stats']['in_progress'] += 1
                    else:
                        data['project_stats']['pending'] += 1

                    # Mark projects for each language/workflow combination
                    for target_lang_name in target_languages:
                        for workflow_key in data['workflow_by_language'].keys():
                            # Check if this workflow_key matches this language
                            if workflow_key.endswith(f" - {target_lang_name}"):
                                data['workflow_by_language'][workflow_key]['projects'] = data['workflow_by_language'][workflow_key].get('projects', 0) + 1

                    # Store project summary
                    data['projects'].append({
                        'id': project_id,
                        'name': project.get('name', 'Unknown'),
                        'status': status,
                        'source_lang': source_lang,
                        'target_langs': ', '.join(target_languages),
                        'total_words': project_total_words,
                        'created_date': None  # Filtering by finish date, not project creation
                    })

        logger.info(f"Aggregated data for {data['project_stats']['total']} projects")
        return data

    def _get_cache_path(self, month: str) -> Path:
        """Get the JSON cache file path for a month (stored locally, not on OneDrive)."""
        cache_dir = Path(__file__).parent / ".cache"
        cache_dir.mkdir(exist_ok=True)
        return cache_dir / f"monthly_{month}.json"

    def _save_month_cache(self, month: str, month_data: Dict):
        """Save processed monthly data to JSON cache."""
        cache_path = self._get_cache_path(month)
        try:
            cache = {
                'languages': {},
                'users': {}
            }
            for wk, metrics in month_data.get('workflow_by_language', {}).items():
                lang = metrics['language']
                if lang not in cache['languages']:
                    cache['languages'][lang] = 0
                cache['languages'][lang] += metrics['words_done']

            for uk, ud in month_data.get('user_statistics', {}).items():
                cache['users'][uk] = {
                    'user': ud['user'],
                    'username': ud.get('username', ''),
                    'language': ud['language'],
                    'total': sum(ud['workflow_steps'].values())
                }

            with open(cache_path, 'w') as f:
                json.dump(cache, f)
            logger.info(f"Cached month data to {cache_path.name}")
        except Exception as e:
            logger.warning(f"Failed to save cache for {month}: {e}")

    def _load_month_cache(self, month: str) -> Dict:
        """Load cached monthly data. Returns None if not available."""
        cache_path = self._get_cache_path(month)
        if not cache_path.exists():
            return None
        try:
            with open(cache_path) as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"Failed to load cache for {month}: {e}")
            return None

    def _get_ytd_language_set(self) -> set:
        """Return the set of language names that had work at any point this year,
        read from the monthly JSON caches (January through the current month).

        Used to keep a consistent language list across reports: a language that
        appeared earlier in the year still shows (as a zero row) even when it had
        no work in the current period. Weekly reports don't compute YTD data, so
        they rely on this cache-based set."""
        langs = set()
        year = self.report_date.year
        for m in range(1, self.report_date.month + 1):
            cached = self._load_month_cache(f"{year}-{m:02d}")
            if cached:
                langs.update(cached.get('languages', {}).keys())
        return langs

    def _report_language_set(self, ytd_monthly_breakdown: Dict) -> set:
        """Full set of languages the language table should list, including ones
        with no work this period (shown as zeros). This is the union of:
          - languages seen this year (YTD breakdown for monthly reports; the
            monthly caches for weekly reports), and
          - every target language any roster volunteer is assigned to.
        The roster part keeps the language report in sync with the user report,
        which lists volunteers under their assigned languages even when idle."""
        if self.weekly:
            langs = self._get_ytd_language_set()
        else:
            langs = set(ytd_monthly_breakdown.get('languages', {}).keys())
        for info in self.get_volunteers().values():
            langs.update(info.get('languages', []))
        return langs

    # ------------------------------------------------------------------
    # Per-project statistics snapshots
    #
    # When a project is archived, the XTM /projects/{id}/statistics endpoint
    # stops returning the per-user breakdown, so the report can only fall back
    # to /metrics (no user attribution) and buckets that work under the generic
    # "Archived User". To avoid that, snapshot_active_projects() periodically
    # caches each active project's raw /statistics (+ /status step finish dates)
    # to .cache/snapshots/. When aggregation later finds a project's live
    # statistics empty, it restores real per-user data from the snapshot.
    # ------------------------------------------------------------------
    def _get_snapshot_dir(self) -> Path:
        """Directory holding per-project raw statistics snapshots."""
        snap_dir = Path(__file__).parent / ".cache" / "snapshots"
        snap_dir.mkdir(parents=True, exist_ok=True)
        return snap_dir

    def _get_snapshot_path(self, project_id: int) -> Path:
        return self._get_snapshot_dir() / f"project_{project_id}.json"

    def _save_project_snapshot(self, project_id, name, status, statistics, status_steps) -> bool:
        """Persist a project's raw /statistics and /status so per-user detail
        survives archival. Only writes when statistics are non-empty, so an
        already-archived project's empty response never overwrites a good snapshot."""
        if not statistics:
            return False
        try:
            payload = {
                'project_id': project_id,
                'name': name,
                'status': status,
                'snapshot_date': datetime.now().isoformat(),
                'statistics': statistics,
                'status_steps': status_steps or {},
            }
            with open(self._get_snapshot_path(project_id), 'w') as f:
                json.dump(payload, f)
            return True
        except Exception as e:
            logger.warning(f"Failed to snapshot project {project_id}: {e}")
            return False

    def _load_project_snapshot(self, project_id: int) -> Dict:
        """Load a project's statistics snapshot. Returns None if not available."""
        path = self._get_snapshot_path(project_id)
        if not path.exists():
            return None
        try:
            with open(path) as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"Failed to load snapshot for project {project_id}: {e}")
            return None

    def _restore_stats_from_snapshot(self, project_id: int):
        """Return (filtered_statistics, status_steps) from a project's snapshot,
        re-applying the current excluded-user filter. Returns ([], {}) if no
        usable snapshot exists."""
        snap = self._load_project_snapshot(project_id)
        if not snap:
            return [], {}
        filtered = self._filter_excluded_from_stats(snap.get('statistics', []))
        if not filtered:
            return [], {}
        return filtered, snap.get('status_steps') or {}

    def snapshot_active_projects(self) -> int:
        """Fetch and cache raw per-user statistics for every project that still
        exposes them (i.e. not yet archived). Run this regularly (e.g. daily) so
        that when a project is later archived the report can restore real user
        names from the snapshot instead of bucketing work under 'Archived User'.
        Returns the number of projects snapshotted."""
        from concurrent.futures import ThreadPoolExecutor, as_completed

        projects = self.get_projects()
        project_map = {p.get('id'): p for p in projects if p.get('id')}
        project_ids = list(project_map.keys())
        logger.info(f"Snapshotting per-user statistics for {len(project_ids)} projects...")

        def fetch(pid):
            try:
                return (pid,
                        self.get_project_statistics_raw(pid),
                        self.get_project_status_with_steps(pid))
            except Exception as e:
                logger.warning(f"Snapshot fetch failed for project {pid}: {e}")
                return pid, [], {}

        saved = 0
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = {executor.submit(fetch, pid): pid for pid in project_ids}
            for future in as_completed(futures):
                pid, stats, status_steps = future.result()
                project = project_map.get(pid, {})
                if self._save_project_snapshot(pid, project.get('name', 'Unknown'),
                                               project.get('status', 'UNKNOWN'),
                                               stats, status_steps):
                    saved += 1

        logger.info(f"Snapshot complete: saved {saved} of {len(project_ids)} projects "
                    f"to {self._get_snapshot_dir()}")
        return saved

    def aggregate_weekly_data(self, start_date: datetime, end_date: datetime) -> Dict:
        """
        Aggregate data for a specific weekly date range.
        Similar to aggregate_monthly_data but uses exact dates instead of month ranges.
        """
        logger.info(f"Aggregating weekly data from {start_date.date()} to {end_date.date()}")

        # Initialize data structures
        data = {
            'project_stats': {
                'total': 0,
                'completed': 0,
                'in_progress': 0,
                'pending': 0
            },
            'workflow_by_language': {},
            'user_statistics': {},
            'projects': []
        }

        # Only fetch projects modified since the week start (see aggregate_monthly_data).
        projects = self.get_projects(modified_from=start_date.strftime('%Y-%m-%d'))

        # Fetch statistics and status for all projects in parallel
        from concurrent.futures import ThreadPoolExecutor, as_completed

        def fetch_project_data(project_id):
            try:
                stats = self.get_project_statistics(project_id)
                status = self.get_project_status_with_steps(project_id)
                return project_id, stats, status
            except Exception as e:
                logger.warning(f"Failed to get data for project {project_id}: {e}")
                return project_id, [], {}

        project_ids = [p.get('id') for p in projects if p.get('id')]
        project_map = {p.get('id'): p for p in projects if p.get('id')}
        project_results = {}

        logger.info(f"Fetching data for {len(project_ids)} projects using parallel requests...")
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = {executor.submit(fetch_project_data, pid): pid for pid in project_ids}
            for future in as_completed(futures):
                pid, stats_list, project_status = future.result()
                project_results[pid] = (stats_list, project_status)

        # Process projects (same logic as monthly)
        for project_id in project_ids:
            project = project_map[project_id]
            stats_list, project_status = project_results[project_id]

            # Restore per-user data from a snapshot if the project is archived
            # (see aggregate_monthly_data for the rationale).
            if not stats_list:
                snap_stats, snap_status = self._restore_stats_from_snapshot(project_id)
                if snap_stats:
                    stats_list = snap_stats
                    if not project_status:
                        project_status = snap_status
                    logger.info(f"Project {project_id} archived; restored per-user "
                                f"statistics from snapshot")

            job_step_dates = {}
            if project_status:
                for job in project_status.get('jobs', []):
                    job_id = job.get('jobId')
                    if job_id:
                        job_step_dates[job_id] = {}
                        for step in job.get('steps', []):
                            step_name = step.get('workflowStepName', '')
                            finish_date = step.get('finishDate')
                            if finish_date and step_name:
                                job_step_dates[job_id][step_name.lower()] = finish_date

            # Check if statistics are empty (archived project)
            if not stats_list or len(stats_list) == 0:
                logger.info(f"Project {project_id} has no statistics (likely archived), using metrics fallback")

                metrics = self.get_project_metrics(project_id)

                if metrics and len(metrics) > 0:
                    project_total_words = 0
                    project_has_work_in_period = False
                    target_languages = []

                    for metric in metrics:
                        target_lang_code = metric.get('targetLanguage', 'unknown')
                        target_lang_name = self._locale_to_language_name(target_lang_code)

                        metrics_progress = metric.get('metricsProgress', {})

                        for step_name, step_metrics in metrics_progress.items():
                            words_done = step_metrics.get('wordsDone', 0)

                            if words_done > 0:
                                clean_step_name = ''.join([c for c in step_name if not c.isdigit()]).strip().lower()

                                # Attribute the cumulative wordsDone to the single period in
                                # which the step was fully completed (latest finish date
                                # across the language's jobs), rather than re-counting it for
                                # every job that finished in the window. See the matching note
                                # in aggregate_monthly_data().
                                step_completed_in_period = False
                                step_finish_dates = []
                                if project_status:
                                    for job in project_status.get('jobs', []):
                                        if job.get('targetLanguage') == target_lang_code:
                                            for step in job.get('steps', []):
                                                if step.get('workflowStepName', '').lower() == step_name.lower():
                                                    finish_date_ts = step.get('finishDate')
                                                    if finish_date_ts:
                                                        step_finish_dates.append(finish_date_ts)

                                if step_finish_dates:
                                    completion_date = datetime.fromtimestamp(max(step_finish_dates) / 1000)
                                    if start_date <= completion_date <= end_date:
                                        step_completed_in_period = True

                                if step_completed_in_period:
                                    project_has_work_in_period = True

                                    if target_lang_name not in target_languages:
                                        target_languages.append(target_lang_name)

                                    # Qualify the generic archived label by language
                                    username = f"Archived User ({target_lang_name})"

                                    user_key = f"{username}|{target_lang_name}"
                                    if user_key not in data['user_statistics']:
                                        data['user_statistics'][user_key] = {
                                            'user': username,
                                            'language': target_lang_name,
                                            'workflow_steps': {}
                                        }

                                    if clean_step_name not in data['user_statistics'][user_key]['workflow_steps']:
                                        data['user_statistics'][user_key]['workflow_steps'][clean_step_name] = 0
                                    data['user_statistics'][user_key]['workflow_steps'][clean_step_name] += words_done

                                    workflow_key = f"{clean_step_name} - {target_lang_name}"

                                    if workflow_key not in data['workflow_by_language']:
                                        data['workflow_by_language'][workflow_key] = {
                                            'workflow_step': clean_step_name,
                                            'language': target_lang_name,
                                            'words_done': 0,
                                            'words_to_be_done': 0,
                                            'projects': 0
                                        }

                                    data['workflow_by_language'][workflow_key]['words_done'] += words_done
                                    project_total_words += words_done

                    if project_has_work_in_period and project_total_words > 0:
                        data['project_stats']['total'] += 1
                        status = project.get('status', 'UNKNOWN')

                        if status == 'FINISHED':
                            data['project_stats']['completed'] += 1
                        elif status in ['IN_PROGRESS', 'STARTED']:
                            data['project_stats']['in_progress'] += 1
                        else:
                            data['project_stats']['pending'] += 1

                        for target_lang_name in target_languages:
                            for workflow_key in data['workflow_by_language'].keys():
                                if workflow_key.endswith(f" - {target_lang_name}"):
                                    data['workflow_by_language'][workflow_key]['projects'] = data['workflow_by_language'][workflow_key].get('projects', 0) + 1

                        data['projects'].append({
                            'id': project_id,
                            'name': project.get('name', 'Unknown'),
                            'status': status,
                            'source_lang': 'en_US',
                            'target_langs': ', '.join(target_languages),
                            'total_words': project_total_words,
                            'created_date': None
                        })

                        logger.info(f"Added archived project {project_id} with {project_total_words} words from metrics")

                continue

            if isinstance(stats_list, list) and stats_list:
                project_total_words = 0
                project_has_work_in_period = False
                target_languages = []

                for lang_stats in stats_list:
                    target_lang_code = lang_stats.get('targetLanguage', 'unknown')
                    target_lang_name = self._locale_to_language_name(target_lang_code)
                    users_statistics = lang_stats.get('usersStatistics', [])

                    for user_stat in users_statistics:
                        steps_statistics = user_stat.get('stepsStatistics', [])

                        for step_stat in steps_statistics:
                            step_name = step_stat.get('workflowStepName', '')
                            # Remove numbers and normalize to lowercase
                            clean_step_name = ''.join([c for c in step_name if not c.isdigit()]).strip().lower()
                            jobs_statistics = step_stat.get('jobsStatistics', [])

                            for job_stat in jobs_statistics:
                                include_this_job = False
                                job_id = job_stat.get('jobId')

                                finish_date_ts = None
                                if job_id and job_id in job_step_dates:
                                    finish_date_ts = job_step_dates[job_id].get(step_name.lower())

                                # Check if completed in weekly date range
                                if finish_date_ts:
                                    completion_date = datetime.fromtimestamp(finish_date_ts / 1000)
                                    # Include end_date in range (<=)
                                    if start_date <= completion_date <= end_date:
                                        include_this_job = True

                                if include_this_job:
                                    source_stats = job_stat.get('sourceStatistics', {})
                                    total_words = source_stats.get('totalWords', 0)

                                    if total_words > 0:
                                        project_has_work_in_period = True
                                        if target_lang_name not in target_languages:
                                            target_languages.append(target_lang_name)

                                        username = self._resolve_user_name(user_stat)
                                        user_key = f"{username}|{target_lang_name}"

                                        if user_key not in data['user_statistics']:
                                            data['user_statistics'][user_key] = {
                                                'user': username,
                                                'username': user_stat.get('username', ''),
                                                'language': target_lang_name,
                                                'workflow_steps': {}
                                            }

                                        if clean_step_name not in data['user_statistics'][user_key]['workflow_steps']:
                                            data['user_statistics'][user_key]['workflow_steps'][clean_step_name] = 0

                                        data['user_statistics'][user_key]['workflow_steps'][clean_step_name] += total_words

                                        # Track by language and workflow step
                                        key = f"{clean_step_name} - {target_lang_name}"
                                        if key not in data['workflow_by_language']:
                                            data['workflow_by_language'][key] = {
                                                'workflow_step': clean_step_name,
                                                'language': target_lang_name,
                                                'words_done': 0,
                                                'words_to_be_done': 0,
                                                'projects': 0
                                            }

                                        data['workflow_by_language'][key]['words_done'] += total_words
                                        project_total_words += total_words

                if project_has_work_in_period:
                    data['project_stats']['total'] += 1
                    data['projects'].append({
                        'id': project_id,
                        'name': project.get('name', 'Unknown'),
                        'total_words': project_total_words,
                        'target_languages': target_languages
                    })

        logger.info(f"Weekly aggregation complete: {data['project_stats']['total']} projects with work in period")
        return data

    def aggregate_ytd_data(self, start_month: str, end_month: str, current_month_data: Dict = None) -> Dict:
        """
        Aggregate YTD data. Uses JSON cache for past months (fast),
        reuses current_month_data if provided (avoids re-querying).
        """
        logger.info(f"Aggregating YTD data from {start_month} to {end_month}")

        start_year, start_month_num = map(int, start_month.split('-'))
        end_year, end_month_num = map(int, end_month.split('-'))

        months = []
        year, month = start_year, start_month_num
        while (year < end_year) or (year == end_year and month <= end_month_num):
            months.append(f"{year}-{month:02d}")
            month += 1
            if month > 12:
                month = 1
                year += 1

        logger.info(f"Processing months: {months}")

        languages = {}
        users = {}

        for month in months:
            # If this is the current month and we already have data, reuse it
            if current_month_data and month == self.report_month:
                logger.info(f"Reusing already-queried data for {month}")
                for wk, metrics in current_month_data['workflow_by_language'].items():
                    lang = metrics['language']
                    if lang not in languages:
                        languages[lang] = {}
                    if month not in languages[lang]:
                        languages[lang][month] = 0
                    languages[lang][month] += metrics['words_done']

                for uk, ud in current_month_data['user_statistics'].items():
                    if uk not in users:
                        users[uk] = {'user': ud['user'], 'username': ud.get('username', ''), 'language': ud['language'], 'months': {}}
                    users[uk]['months'][month] = sum(ud['workflow_steps'].values())
                continue

            # Try JSON cache first (has proper names from previous API queries)
            cached = self._load_month_cache(month)
            if cached:
                logger.info(f"Using cached data for {month}")
                for lang, total in cached['languages'].items():
                    if lang not in languages:
                        languages[lang] = {}
                    languages[lang][month] = total

                for uk, ud in cached['users'].items():
                    if uk not in users:
                        users[uk] = {'user': ud['user'], 'username': ud.get('username', ''), 'language': ud['language'], 'months': {}}
                    users[uk]['months'][month] = ud['total']
                continue

            # No cache — query API
            logger.info(f"No cache for {month}, querying API...")
            try:
                month_data = self.aggregate_monthly_data(month, month)

                for wk, metrics in month_data['workflow_by_language'].items():
                    lang = metrics['language']
                    if lang not in languages:
                        languages[lang] = {}
                    if month not in languages[lang]:
                        languages[lang][month] = 0
                    languages[lang][month] += metrics['words_done']

                for uk, ud in month_data['user_statistics'].items():
                    if uk not in users:
                        users[uk] = {'user': ud['user'], 'username': ud.get('username', ''), 'language': ud['language'], 'months': {}}
                    users[uk]['months'][month] = sum(ud['workflow_steps'].values())

                # Save to cache for future runs
                self._save_month_cache(month, month_data)

            except Exception as e:
                logger.error(f"API query failed for {month}: {e}", exc_info=True)

        return {
            'months': months,
            'languages': languages,
            'users': users
        }

    @staticmethod
    def _generate_bar_chart_base64(labels, datasets, title, stacked=False, width=10, height=5) -> str:
        """Generate a bar chart as base64-encoded PNG."""
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import numpy as np
        import base64
        from io import BytesIO

        fig, ax = plt.subplots(figsize=(width, height))
        x = np.arange(len(labels))
        bar_width = 0.6

        if stacked and len(datasets) > 1:
            bottom = np.zeros(len(labels))
            for ds in datasets:
                ax.bar(x, ds['data'], bar_width, label=ds['label'],
                       color=ds.get('backgroundColor', '#36A2EB'), bottom=bottom)
                bottom += np.array(ds['data'])
            ax.legend(fontsize=8)
        else:
            data = datasets[0]['data'] if datasets else []
            color = datasets[0].get('backgroundColor', '#36A2EB') if datasets else '#36A2EB'
            ax.bar(x, data, bar_width, color=color)

        ax.set_title(title, fontsize=12, fontweight='bold')
        ax.set_xticks(x)
        # Truncate long labels
        short_labels = [l[:20] + '...' if len(l) > 20 else l for l in labels]
        ax.set_xticklabels(short_labels, rotation=45, ha='right', fontsize=7)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f'{int(v):,}'))
        ax.grid(axis='y', alpha=0.3)
        plt.tight_layout()

        buf = BytesIO()
        fig.savefig(buf, format='png', dpi=120, bbox_inches='tight')
        plt.close(fig)
        buf.seek(0)
        return base64.b64encode(buf.read()).decode('utf-8')

    @staticmethod
    def _generate_line_chart_base64(months, datasets, title, max_series=10, width=10, height=5) -> str:
        """Generate a line chart as base64-encoded PNG."""
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import base64
        from io import BytesIO

        fig, ax = plt.subplots(figsize=(width, height))
        colors = ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF',
                  '#FF9F40', '#E7E9ED', '#66BB6A', '#AB47BC', '#FF7043']

        for idx, ds in enumerate(datasets[:max_series]):
            ax.plot(months, ds['data'], marker='o', markersize=4,
                    label=ds['label'][:25], color=colors[idx % len(colors)], linewidth=2)

        ax.set_title(title, fontsize=12, fontweight='bold')
        ax.set_xlabel('Month', fontsize=9)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f'{int(v):,}'))
        ax.legend(fontsize=7, loc='upper left', bbox_to_anchor=(1.02, 1))
        ax.grid(alpha=0.3)
        plt.tight_layout()

        buf = BytesIO()
        fig.savefig(buf, format='png', dpi=120, bbox_inches='tight')
        plt.close(fig)
        buf.seek(0)
        return base64.b64encode(buf.read()).decode('utf-8')

    def create_excel_report(self, monthly_data: Dict, ytd_monthly_breakdown: Dict, output_path: str) -> str:
        """Create an Excel report with monthly and YTD data."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, LineChart, Reference

        logger.info(f"Creating Excel report at {output_path}")

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # === CURRENT PERIOD SHEET ===
        self._create_monthly_sheet(wb, monthly_data, ytd_monthly_breakdown)

        # === YTD SHEET (only for monthly reports) ===
        if not self.weekly:
            self._create_ytd_sheet(wb, ytd_monthly_breakdown)

        # === USER STATISTICS SHEETS ===
        self._create_user_monthly_sheet(wb, monthly_data)
        if not self.weekly:
            self._create_user_ytd_sheet(wb, ytd_monthly_breakdown)

        # === VOLUNTEER HOURS SHEET (from XTM login/logout history) ===
        self._create_volunteer_hours_sheet(wb)

        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel report saved to {output_path}")
        return output_path

    def _create_monthly_sheet(self, wb: 'Workbook', monthly_data: Dict, ytd_monthly_breakdown: Dict = None):
        """Create the current period data sheet with workflow breakdown."""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference

        sheet_name = f"Weekly - {self.report_week_label.replace('Week of ', '')}" if self.weekly else f"Monthly - {self.report_month}"
        ws = wb.create_sheet(sheet_name)

        # Extract data
        workflow_by_language = monthly_data.get('workflow_by_language', {})

        # Organize by language
        monthly_languages = {}
        for workflow_key, metrics in workflow_by_language.items():
            language = metrics['language']
            workflow_step = metrics['workflow_step']
            words = metrics['words_done']

            if language not in monthly_languages:
                monthly_languages[language] = {}
            monthly_languages[language][workflow_step] = words

        # Keep every language seen this year, even with no work this period
        # (shows as a zero row), matching the HTML report.
        for lang in self._report_language_set(ytd_monthly_breakdown or {}):
            monthly_languages.setdefault(lang, {})

        # Get all workflow steps
        all_steps = set()
        for lang_steps in monthly_languages.values():
            all_steps.update(lang_steps.keys())
        workflow_steps = sorted(all_steps, key=lambda x: ['translate', 'correct', 'final review'].index(x) if x in ['translate', 'correct', 'final review'] else 999)

        # Calculate totals
        monthly_language_totals = {lang: sum(steps.values()) for lang, steps in monthly_languages.items()}

        # Write headers
        headers = ['Language'] + workflow_steps + ['Total']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Write data rows
        sorted_languages = sorted(monthly_language_totals.keys(), key=lambda x: monthly_language_totals[x], reverse=True)
        for row_idx, language in enumerate(sorted_languages, 2):
            ws.cell(row=row_idx, column=1, value=language)

            for col_idx, step in enumerate(workflow_steps, 2):
                words = monthly_languages[language].get(step, 0)
                cell = ws.cell(row=row_idx, column=col_idx, value=words)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')

            # Total column
            total_cell = ws.cell(row=row_idx, column=len(headers), value=monthly_language_totals[language])
            total_cell.number_format = '#,##0'
            total_cell.font = Font(bold=True)
            total_cell.fill = PatternFill(start_color="F0F7FF", end_color="F0F7FF", fill_type="solid")
            total_cell.alignment = Alignment(horizontal='right')

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        # Enable AutoFilter
        ws.auto_filter.ref = ws.dimensions

        # Add bar chart — only languages with work this period. They're sorted
        # to the top, so capping the row range excludes the zero-filled ones
        # (which remain in the table) without cluttering the chart.
        chart_lang_count = sum(1 for lang in sorted_languages if monthly_language_totals[lang] > 0)
        if chart_lang_count > 0:
            chart = BarChart()
            chart.title = "Words Processed by Language"
            chart.y_axis.title = "Words"
            chart.x_axis.title = "Language"

            # Total column data
            data = Reference(ws, min_col=len(headers), min_row=1, max_row=chart_lang_count + 1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=chart_lang_count + 1)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 15
            chart.width = 25

            ws.add_chart(chart, f"A{len(sorted_languages) + 3}")

    def _create_ytd_sheet(self, wb: 'Workbook', ytd_monthly_breakdown: Dict):
        """Create the YTD sheet with monthly breakdown."""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import LineChart, Reference

        ws = wb.create_sheet(f"YTD - {self.ytd_start_month} to {self.ytd_end_month}")

        months = ytd_monthly_breakdown['months']
        ytd_languages = ytd_monthly_breakdown['languages']

        # Write headers
        headers = ['Language'] + months + ['Total']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Write data rows
        sorted_languages = sorted(ytd_languages.items(), key=lambda x: sum(x[1].values()), reverse=True)
        for row_idx, (language, month_data) in enumerate(sorted_languages, 2):
            ws.cell(row=row_idx, column=1, value=language)

            row_total = 0
            for col_idx, month in enumerate(months, 2):
                words = month_data.get(month, 0)
                cell = ws.cell(row=row_idx, column=col_idx, value=words)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
                row_total += words

            # Total column
            total_cell = ws.cell(row=row_idx, column=len(headers), value=row_total)
            total_cell.number_format = '#,##0'
            total_cell.font = Font(bold=True)
            total_cell.fill = PatternFill(start_color="F0F7FF", end_color="F0F7FF", fill_type="solid")
            total_cell.alignment = Alignment(horizontal='right')

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        # Enable AutoFilter
        ws.auto_filter.ref = ws.dimensions

        # Add line chart for top 10 languages — only languages with work this year.
        # They're sorted to the top, so capping the range excludes zero rows.
        nonzero_langs = sum(1 for _, md in sorted_languages if sum(md.values()) > 0)
        if nonzero_langs > 0:
            chart = LineChart()
            chart.title = "Language Translation Trends (Top 10)"
            chart.y_axis.title = "Words"
            chart.x_axis.title = "Month"

            # Add data for top 10 languages
            num_langs = min(10, nonzero_langs)
            for row_idx in range(2, num_langs + 2):
                data = Reference(ws, min_col=2, max_col=len(months) + 1, min_row=row_idx, max_row=row_idx)
                chart.add_data(data, titles_from_data=False)

            # Set categories (months)
            cats = Reference(ws, min_col=2, max_col=len(months) + 1, min_row=1, max_row=1)
            chart.set_categories(cats)

            # Series labels are automatically pulled from the language column (col 1)
            # We need to add titles manually using SeriesLabel
            from openpyxl.chart.series import SeriesLabel
            from openpyxl.chart.text import StrRef, Text
            for idx, (language, _) in enumerate(sorted_languages[:num_langs]):
                # Create a simple text-based title
                text = Text()
                text.rich = None
                text.plain = language
                series_label = SeriesLabel()
                series_label.strRef = None
                series_label.v = language
                chart.series[idx].tx = series_label

            chart.height = 15
            chart.width = 25

            ws.add_chart(chart, f"A{len(sorted_languages) + 3}")

    def _create_user_monthly_sheet(self, wb: 'Workbook', monthly_data: Dict):
        """Create the monthly user statistics sheet."""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference

        user_statistics = monthly_data.get('user_statistics', {})
        if not user_statistics:
            return  # Skip if no user data

        sheet_name = f"User Stats - {self.report_week_label.replace('Week of ', '')}" if self.weekly else f"User Stats - {self.report_month}"
        ws = wb.create_sheet(sheet_name)

        # Get all workflow steps
        all_steps = set()
        for user_data in user_statistics.values():
            all_steps.update(user_data['workflow_steps'].keys())
        workflow_steps = sorted(all_steps, key=lambda x: ['translate', 'correct', 'final review'].index(x) if x in ['translate', 'correct', 'final review'] else 999)

        # Write headers
        headers = ['User', 'Language'] + workflow_steps + ['Total']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Write data rows - sort by total words descending
        sorted_users = sorted(user_statistics.items(),
                             key=lambda x: sum(x[1]['workflow_steps'].values()),
                             reverse=True)

        for row_idx, (user_key, user_data) in enumerate(sorted_users, 2):
            ws.cell(row=row_idx, column=1, value=user_data['user'])
            ws.cell(row=row_idx, column=2, value=user_data['language'])

            row_total = 0
            for col_idx, step in enumerate(workflow_steps, 3):
                words = user_data['workflow_steps'].get(step, 0)
                cell = ws.cell(row=row_idx, column=col_idx, value=words)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
                row_total += words

            # Total column
            total_cell = ws.cell(row=row_idx, column=len(headers), value=row_total)
            total_cell.number_format = '#,##0'
            total_cell.font = Font(bold=True)
            total_cell.fill = PatternFill(start_color="F0F7FF", end_color="F0F7FF", fill_type="solid")
            total_cell.alignment = Alignment(horizontal='right')

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        # Enable AutoFilter
        ws.auto_filter.ref = ws.dimensions

        # Add bar chart for top 20 users — only volunteers with work this period.
        # They're sorted to the top, so capping the range excludes zero rows.
        nonzero_users = sum(1 for _, ud in sorted_users if sum(ud['workflow_steps'].values()) > 0)
        if nonzero_users > 0:
            chart = BarChart()
            chart.title = "Words Processed by User (Top 20)"
            chart.y_axis.title = "Words"
            chart.x_axis.title = "User"

            # Total column data (limit to top 20)
            num_users = min(20, nonzero_users)
            data = Reference(ws, min_col=len(headers), min_row=1, max_row=num_users + 1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=num_users + 1)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 15
            chart.width = 25

            ws.add_chart(chart, f"A{len(sorted_users) + 3}")

    def _create_user_ytd_sheet(self, wb: 'Workbook', ytd_monthly_breakdown: Dict):
        """Create the YTD user statistics sheet."""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.series import SeriesLabel

        months = ytd_monthly_breakdown['months']
        ytd_users = ytd_monthly_breakdown.get('users', {})

        if not ytd_users:
            return  # Skip if no user data

        ws = wb.create_sheet(f"User Stats - YTD")

        # Write headers
        headers = ['User', 'Language'] + months + ['Total']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Write data rows - sort by total words descending
        sorted_users = sorted(ytd_users.items(),
                             key=lambda x: sum(x[1]['months'].values()),
                             reverse=True)

        for row_idx, (user_key, user_data) in enumerate(sorted_users, 2):
            ws.cell(row=row_idx, column=1, value=user_data['user'])
            ws.cell(row=row_idx, column=2, value=user_data['language'])

            row_total = 0
            for col_idx, month in enumerate(months, 3):
                words = user_data['months'].get(month, 0)
                cell = ws.cell(row=row_idx, column=col_idx, value=words)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
                row_total += words

            # Total column
            total_cell = ws.cell(row=row_idx, column=len(headers), value=row_total)
            total_cell.number_format = '#,##0'
            total_cell.font = Font(bold=True)
            total_cell.fill = PatternFill(start_color="F0F7FF", end_color="F0F7FF", fill_type="solid")
            total_cell.alignment = Alignment(horizontal='right')

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        # Enable AutoFilter
        ws.auto_filter.ref = ws.dimensions

        # Add line chart for top 10 users — only volunteers with work this year.
        # They're sorted to the top, so capping the range excludes zero rows.
        nonzero_users = sum(1 for _, ud in sorted_users if sum(ud['months'].values()) > 0)
        if nonzero_users > 0:
            chart = LineChart()
            chart.title = "User Productivity Trends (Top 10)"
            chart.y_axis.title = "Words"
            chart.x_axis.title = "Month"

            # Add data for top 10 users
            num_users = min(10, nonzero_users)
            for row_idx in range(2, num_users + 2):
                data = Reference(ws, min_col=3, max_col=len(months) + 2, min_row=row_idx, max_row=row_idx)
                chart.add_data(data, titles_from_data=False)

            # Set categories (months)
            cats = Reference(ws, min_col=3, max_col=len(months) + 2, min_row=1, max_row=1)
            chart.set_categories(cats)

            # Add series labels using SeriesLabel
            for idx, (user_key, user_data) in enumerate(sorted_users[:num_users]):
                username = user_data['user']
                language = user_data['language']
                series_label = SeriesLabel()
                series_label.strRef = None
                series_label.v = f"{username} ({language})"
                chart.series[idx].tx = series_label

            chart.height = 15
            chart.width = 25

            ws.add_chart(chart, f"A{len(sorted_users) + 3}")

    def create_combined_html_report(self, monthly_data: Dict, ytd_monthly_breakdown: Dict, output_path: str) -> str:
        """Create a combined interactive HTML report with both monthly and YTD data."""
        import json

        # === MONTHLY DATA ===
        workflow_by_language = monthly_data.get('workflow_by_language', {})
        user_statistics = monthly_data.get('user_statistics', {})

        # Organize monthly data by language and workflow step
        monthly_languages = {}
        for workflow_key, metrics in workflow_by_language.items():
            language = metrics['language']
            workflow_step = metrics['workflow_step']
            words = metrics['words_done']

            if language not in monthly_languages:
                monthly_languages[language] = {}
            monthly_languages[language][workflow_step] = words

        # Keep every language seen this year in the current-period table/chart,
        # even when it had no work this period — it shows as a zero row / bar
        # rather than disappearing.
        for lang in self._report_language_set(ytd_monthly_breakdown):
            monthly_languages.setdefault(lang, {})

        # Calculate totals per language for monthly
        monthly_language_totals = {lang: sum(steps.values()) for lang, steps in monthly_languages.items()}

        # Get all workflow steps
        all_steps = set()
        for lang_steps in monthly_languages.values():
            all_steps.update(lang_steps.keys())
        workflow_steps = sorted(all_steps, key=lambda x: ['translate', 'correct', 'final review'].index(x) if x in ['translate', 'correct', 'final review'] else 999)

        # Prepare monthly bar chart data (stacked by workflow step).
        # Charts show only languages with work this period; the zero-filled
        # languages stay in the table/filters but would just clutter the chart.
        monthly_lang_labels = [lang for lang in sorted(monthly_language_totals.keys(), key=lambda x: monthly_language_totals[x], reverse=True)
                               if monthly_language_totals[lang] > 0]
        monthly_workflow_datasets = []
        colors = ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF']
        for idx, step in enumerate(workflow_steps):
            data_points = [monthly_languages.get(lang, {}).get(step, 0) for lang in monthly_lang_labels]
            monthly_workflow_datasets.append({
                'label': step,
                'data': data_points,
                'backgroundColor': colors[idx % len(colors)]
            })

        # Monthly user data for chart — only volunteers with work this period;
        # zero-work volunteers stay in the table/filters but would clutter the chart.
        monthly_user_labels = []
        monthly_user_data_points = []
        for user_key, user_data in sorted(user_statistics.items(), key=lambda x: sum(x[1]['workflow_steps'].values()), reverse=True):
            total = sum(user_data['workflow_steps'].values())
            if total <= 0:
                continue
            monthly_user_labels.append(f"{user_data['user']} ({user_data['language']})")
            monthly_user_data_points.append(total)

        # Create monthly language table HTML
        monthly_lang_table_html = ""
        for language in sorted(monthly_language_totals.keys(), key=lambda x: monthly_language_totals[x], reverse=True):
            row = f"<tr><td>{language}</td>"
            for step in workflow_steps:
                row += f"<td class='number'>{monthly_languages[language].get(step, 0):,}</td>"
            row += f"<td class='number total'><strong>{monthly_language_totals[language]:,}</strong></td></tr>"
            monthly_lang_table_html += row

        # Create monthly user table HTML and build user-language mapping
        monthly_user_table_html = ""
        monthly_user_lang_map = {}  # language -> list of users
        for user_key, user_data in sorted(user_statistics.items(), key=lambda x: sum(x[1]['workflow_steps'].values()), reverse=True):
            username = user_data['user']
            language = user_data['language']

            # Build mapping
            if language not in monthly_user_lang_map:
                monthly_user_lang_map[language] = []
            if username not in monthly_user_lang_map[language]:
                monthly_user_lang_map[language].append(username)

            row = f"<tr><td>{username}</td><td>{language}</td>"
            for step in workflow_steps:
                row += f"<td class='number'>{user_data['workflow_steps'].get(step, 0):,}</td>"
            total = sum(user_data['workflow_steps'].values())
            row += f"<td class='number total'><strong>{total:,}</strong></td></tr>"
            monthly_user_table_html += row

        # === YTD DATA (only for monthly reports) ===
        if not self.weekly:
            months = ytd_monthly_breakdown['months']
            ytd_languages = ytd_monthly_breakdown['languages']
            ytd_users = ytd_monthly_breakdown.get('users', {})

            # Prepare YTD line chart data
            ytd_lang_datasets = []
            lang_colors = [
                '#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF',
                '#FF9F40', '#FF6384', '#C9CBCF', '#4BC0C0', '#FF9F40'
            ]

            # Charts only: drop languages with no work this year (they remain in the table).
            sorted_ytd_languages = [item for item in sorted(ytd_languages.items(), key=lambda x: sum(x[1].values()), reverse=True)
                                    if sum(item[1].values()) > 0]

            for idx, (language, month_data) in enumerate(sorted_ytd_languages):
                data_points = [month_data.get(month, 0) for month in months]
                ytd_lang_datasets.append({
                    'label': language,
                    'data': data_points,
                    'borderColor': lang_colors[idx % len(lang_colors)],
                    'backgroundColor': lang_colors[idx % len(lang_colors)] + '20',
                    'tension': 0.4
                })

            # YTD user data
            ytd_user_datasets = []
            # Charts only: drop zero-work volunteers (they remain in the table below).
            sorted_ytd_users = [item for item in sorted(ytd_users.items(), key=lambda x: sum(x[1]['months'].values()), reverse=True)
                                if sum(item[1]['months'].values()) > 0]

            for idx, (user_key, user_data) in enumerate(sorted_ytd_users):
                username = user_data['user']
                language = user_data['language']
                data_points = [user_data['months'].get(month, 0) for month in months]
                ytd_user_datasets.append({
                    'label': f"{username} ({language})",
                    'data': data_points,
                    'borderColor': lang_colors[idx % len(lang_colors)],
                    'backgroundColor': lang_colors[idx % len(lang_colors)] + '20',
                    'tension': 0.4
                })

            # Create YTD language table HTML
            ytd_lang_table_html = ""
            for language, month_data in sorted(ytd_languages.items(), key=lambda x: sum(x[1].values()), reverse=True):
                row = f"<tr><td>{language}</td>"
                for month in months:
                    row += f"<td class='number'>{month_data.get(month, 0):,}</td>"
                row += f"<td class='number total'><strong>{sum(month_data.values()):,}</strong></td></tr>"
                ytd_lang_table_html += row

            # Create YTD user table HTML and build user-language mapping
            ytd_user_table_html = ""
            ytd_user_lang_map = {}  # language -> list of users
            for user_key, user_data in sorted(ytd_users.items(), key=lambda x: sum(x[1]['months'].values()), reverse=True):
                username = user_data['user']
                language = user_data['language']

                # Build mapping
                if language not in ytd_user_lang_map:
                    ytd_user_lang_map[language] = []
                if username not in ytd_user_lang_map[language]:
                    ytd_user_lang_map[language].append(username)

                row = f"<tr><td>{username}</td><td>{language}</td>"
                for month in months:
                    row += f"<td class='number'>{user_data['months'].get(month, 0):,}</td>"
                row += f"<td class='number total'><strong>{sum(user_data['months'].values()):,}</strong></td></tr>"
                ytd_user_table_html += row
        else:
            # Weekly reports don't have YTD data
            months = []
            ytd_languages = {}
            ytd_users = {}
            ytd_lang_datasets = []
            ytd_user_datasets = []
            ytd_lang_table_html = ""
            ytd_user_table_html = ""
            ytd_user_lang_map = {}

        # Generate static chart images
        logger.info("Generating chart images...")
        monthly_lang_chart_b64 = self._generate_bar_chart_base64(
            monthly_lang_labels, monthly_workflow_datasets,
            'Words Processed by Language and Workflow Step', stacked=True)

        monthly_user_ds = [{'label': 'Total Words', 'data': monthly_user_data_points[:20],
                            'backgroundColor': '#36A2EB'}]
        monthly_user_chart_b64 = self._generate_bar_chart_base64(
            [l[:30] for l in monthly_user_labels[:20]], monthly_user_ds,
            'Words Processed by User')

        # YTD charts only for monthly reports
        if not self.weekly:
            ytd_lang_ds = [{'label': l, 'data': [ytd_languages[l].get(m, 0) for m in months]}
                           for l, _ in sorted_ytd_languages[:10]]
            ytd_lang_chart_b64 = self._generate_line_chart_base64(
                months, ytd_lang_ds, 'Language Trends (Top 10)')

            ytd_user_ds = [{'label': f"{ud['user']} ({ud['language']})",
                            'data': [ud['months'].get(m, 0) for m in months]}
                           for _, ud in sorted_ytd_users[:10]]
            ytd_user_chart_b64 = self._generate_line_chart_base64(
                months, ytd_user_ds, 'User Productivity (Top 10)')
        else:
            ytd_lang_chart_b64 = ""
            ytd_user_chart_b64 = ""

        # Generate YTD HTML section (only for monthly reports)
        if not self.weekly:
            ytd_section_html = f"""
        <!-- YTD SECTION -->
        <div class="section-divider">
            <h2>📈 Year-to-Date Report - {self.ytd_start_month} to {self.ytd_end_month}</h2>
        </div>

        <div class="filter-panel">
            <h3>🔍 Filters</h3>
            <div class="filter-group">
                <div class="filter-item">
                    <label>Languages:</label>
                    <input type="text" id="ytdLangSearch" placeholder="Type to search..." onkeyup="filterCheckboxes('ytdLangCheckboxes', this.value)">
                    <div class="checkbox-list" id="ytdLangCheckboxes">
                        {''.join(['<div class="checkbox-item"><input type="checkbox" id="ytdLang_' + lang.replace(" ", "_") + '" value="' + lang + '" onchange="applyFilters(' + "'ytd'" + ')"><label for="ytdLang_' + lang.replace(" ", "_") + '">' + lang + '</label></div>' for lang in sorted(ytd_languages.keys())])}
                    </div>
                </div>
                <div class="filter-item">
                    <label>Users:</label>
                    <input type="text" id="ytdUserSearch" placeholder="Type to search..." onkeyup="filterCheckboxes('ytdUserCheckboxes', this.value)">
                    <div class="checkbox-list" id="ytdUserCheckboxes">
                        {''.join(['<div class="checkbox-item" data-lang="' + user_data["language"] + '"><input type="checkbox" id="ytdUser_' + user_data["user"].replace(" ", "_") + '" value="' + user_data["user"] + '" onchange="applyFilters(' + "'ytd'" + ')"><label for="ytdUser_' + user_data["user"].replace(" ", "_") + '">' + user_data["user"] + '</label></div>' for user_key, user_data in sorted(ytd_users.items(), key=lambda x: x[1]["user"])])}
                    </div>
                </div>
            </div>
        </div>

        <div class="data-row">
            <div class="data-card">
                <h3>🌍 Language Translation Trends</h3>
                <img class="chart-img" id="ytdLangImg" src="data:image/png;base64,{ytd_lang_chart_b64}" alt="YTD Language Chart">
                <div class="chart-container" id="ytdLangChartWrap"><canvas id="ytdLanguageChart"></canvas></div>
            </div>
            <div class="data-card">
                <h3>📊 Language Data</h3>
                <div class="table-container">
                    <table id="ytdLangTable">
                        <thead>
                            <tr>
                                <th class="sortable" onclick="sortTable('ytdLangTable', 0)">Language</th>
                                {''.join([f'<th class="number sortable" onclick="sortTable(' + "'ytdLangTable', " + str(idx+1) + ')">' + month + '</th>' for idx, month in enumerate(months)])}
                                <th class="number sortable" onclick="sortTable('ytdLangTable', {len(months)+1})">Total</th>
                            </tr>
                        </thead>
                        <tbody>
                            {ytd_lang_table_html}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>

        <div class="data-row">
            <div class="data-card">
                <h3>👥 User Productivity Trends</h3>
                <img class="chart-img" id="ytdUserImg" src="data:image/png;base64,{ytd_user_chart_b64}" alt="YTD User Chart">
                <div class="chart-container" id="ytdUserChartWrap"><canvas id="ytdUserChart"></canvas></div>
            </div>
            <div class="data-card">
                <h3>📊 User Data</h3>
                <div class="table-container">
                    <table id="ytdUserTable">
                        <thead>
                            <tr>
                                <th class="sortable" onclick="sortTable('ytdUserTable', 0)">Name</th>
                                <th class="sortable" onclick="sortTable('ytdUserTable', 1)">Language</th>
                                {''.join([f'<th class="number sortable" onclick="sortTable(' + "'ytdUserTable', " + str(idx+2) + ')">' + month + '</th>' for idx, month in enumerate(months)])}
                                <th class="number sortable" onclick="sortTable('ytdUserTable', {len(months)+2})">Total</th>
                            </tr>
                        </thead>
                        <tbody>
                            {ytd_user_table_html}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
"""
        else:
            ytd_section_html = ""

        # Build YTD chart JS separately to avoid nested f-strings (those require Python 3.12+; launchd runs Python 3.9)
        if not self.weekly:
            ytd_charts_js = f"""ytdLanguageChart = new Chart(document.getElementById('ytdLanguageChart'), {{
                type: 'line',
                data: {{ labels: {json.dumps(months)}, datasets: {json.dumps(ytd_lang_datasets)} }},
                options: {{ responsive: true, maintainAspectRatio: false,
                    plugins: {{ title: {{ display: true, text: 'Language Trends' }}, legend: {{ position: 'right', labels: {{ boxWidth: 12, font: {{ size: 9 }} }} }} }},
                    scales: {{ y: {{ beginAtZero: true, ticks: {{ callback: fmtTick }} }} }}
                }}
            }});

            ytdUserChart = new Chart(document.getElementById('ytdUserChart'), {{
                type: 'line',
                data: {{ labels: {json.dumps(months)}, datasets: {json.dumps(ytd_user_datasets)} }},
                options: {{ responsive: true, maintainAspectRatio: false,
                    plugins: {{ title: {{ display: true, text: 'User Productivity' }}, legend: {{ position: 'right', labels: {{ boxWidth: 12, font: {{ size: 9 }} }} }} }},
                    scales: {{ y: {{ beginAtZero: true, ticks: {{ callback: fmtTick }} }} }}
                }}
            }});"""
        else:
            ytd_charts_js = ""

        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>XTM Report - {self.report_month}</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 20px; color: #333; }}
        .container {{ max-width: 1600px; margin: 0 auto; }}
        .header {{ background: white; padding: 30px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); margin-bottom: 20px; }}
        h1 {{ color: #366092; font-size: 2em; margin-bottom: 10px; }}
        .subtitle {{ color: #666; font-size: 1.1em; }}
        .section-divider {{ background: white; padding: 20px 30px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); margin: 30px 0 20px 0; }}
        .section-divider h2 {{ color: #366092; margin: 0; font-size: 1.8em; }}
        .filter-panel {{ background: white; padding: 20px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); margin-bottom: 20px; }}
        .filter-group {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
        .filter-item {{ display: flex; flex-direction: column; }}
        .filter-item label {{ font-weight: 600; color: #366092; margin-bottom: 8px; font-size: 1.1em; }}
        .filter-item input[type="text"] {{ padding: 8px; border: 2px solid #ddd; border-radius: 5px; font-size: 16px; margin-bottom: 10px; }}
        .filter-item input[type="text"]:focus {{ outline: none; border-color: #366092; }}
        .checkbox-list {{ max-height: 200px; overflow-y: auto; border: 1px solid #ddd; border-radius: 5px; padding: 10px; background: #fafafa; }}
        .checkbox-item {{ display: flex; align-items: center; padding: 5px 0; }}
        .checkbox-item input {{ margin-right: 8px; cursor: pointer; }}
        .checkbox-item label {{ cursor: pointer; margin: 0; font-weight: normal; color: #333; }}
        .data-row {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 20px; }}
        .data-card {{ background: white; padding: 20px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); overflow: hidden; }}
        h3 {{ color: #366092; margin-bottom: 15px; padding-bottom: 8px; border-bottom: 3px solid #366092; font-size: 1.2em; }}
        .chart-img {{ width: 100%; height: auto; display: block; }}
        .chart-container {{ position: relative; height: 400px; display: none; }}
        @media (max-width: 768px) {{
            .chart-container {{ height: 250px; }}
        }}
        .table-container {{ overflow-x: auto; max-height: 400px; overflow-y: auto; -webkit-overflow-scrolling: touch; }}
        table {{ width: 100%; border-collapse: collapse; font-size: 0.85em; }}
        th {{ background: #366092; color: white; padding: 8px; text-align: left; font-weight: 600; cursor: pointer; user-select: none; position: sticky; top: 0; z-index: 10; white-space: nowrap; }}
        th:hover {{ background: #2a4d73; }}
        th.sortable:after {{ content: ' ⇅'; opacity: 0.5; font-size: 0.8em; }}
        th.sort-asc:after {{ content: ' ▲'; opacity: 1; }}
        th.sort-desc:after {{ content: ' ▼'; opacity: 1; }}
        td {{ padding: 6px 8px; border-bottom: 1px solid #e0e0e0; white-space: nowrap; }}
        tr:hover {{ background: #f5f5f5; }}
        .number {{ text-align: right; }}
        .total {{ background: #f0f7ff; }}
        @media (max-width: 1400px) {{
            .data-row {{ grid-template-columns: 1fr; }}
            .filter-group {{ grid-template-columns: 1fr; }}
        }}
        @media (max-width: 768px) {{
            body {{ padding: 8px; }}
            .header {{ padding: 15px; }}
            h1 {{ font-size: 1.4em; }}
            .subtitle {{ font-size: 0.9em; }}
            .section-divider {{ padding: 12px 15px; }}
            .section-divider h2 {{ font-size: 1.3em; }}
            .filter-panel {{ padding: 12px; }}
            .data-card {{ padding: 10px; }}
            h3 {{ font-size: 1em; }}
            .checkbox-list {{ max-height: 150px; }}
            table {{ font-size: 0.75em; }}
            th, td {{ padding: 4px 6px; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 XTM Translation Report</h1>
            <div class="subtitle">
                {'<strong>Report Period:</strong> ' + (self.report_period if self.weekly else f"{self.report_month_name} {self.report_month.split('-')[0]}") + '<br>' if (self.weekly or self.report_month_name) else ''}
                {'' if self.weekly else f'<strong>Year-to-Date:</strong> {self.ytd_start_month} to {self.ytd_end_month}<br>'}
                <strong>Generated:</strong> {self.report_date.strftime('%Y-%m-%d %H:%M')}
            </div>
        </div>

        <!-- CURRENT PERIOD SECTION -->
        <div class="section-divider">
            <h2>{'📅 Weekly Report - ' + self.report_week_label if self.weekly else '📅 Monthly Report - ' + self.report_month_name}</h2>
        </div>

        <div class="filter-panel">
            <h3>🔍 Filters</h3>
            <div class="filter-group">
                <div class="filter-item">
                    <label>Languages:</label>
                    <input type="text" id="monthlyLangSearch" placeholder="Type to search..." onkeyup="filterCheckboxes('monthlyLangCheckboxes', this.value)">
                    <div class="checkbox-list" id="monthlyLangCheckboxes">
                        {''.join(['<div class="checkbox-item"><input type="checkbox" id="monthlyLang_' + lang.replace(" ", "_") + '" value="' + lang + '" onchange="applyFilters(' + "'monthly'" + ')"><label for="monthlyLang_' + lang.replace(" ", "_") + '">' + lang + '</label></div>' for lang in sorted(monthly_language_totals.keys())])}
                    </div>
                </div>
                <div class="filter-item">
                    <label>Users:</label>
                    <input type="text" id="monthlyUserSearch" placeholder="Type to search..." onkeyup="filterCheckboxes('monthlyUserCheckboxes', this.value)">
                    <div class="checkbox-list" id="monthlyUserCheckboxes">
                        {''.join(['<div class="checkbox-item" data-lang="' + user_data["language"] + '"><input type="checkbox" id="monthlyUser_' + user_data["user"].replace(" ", "_") + '" value="' + user_data["user"] + '" onchange="applyFilters(' + "'monthly'" + ')"><label for="monthlyUser_' + user_data["user"].replace(" ", "_") + '">' + user_data["user"] + '</label></div>' for user_key, user_data in sorted(user_statistics.items(), key=lambda x: x[1]["user"])])}
                    </div>
                </div>
            </div>
        </div>

        <div class="data-row">
            <div class="data-card">
                <h3>🌍 Language Translation Volume</h3>
                <img class="chart-img" id="monthlyLangImg" src="data:image/png;base64,{monthly_lang_chart_b64}" alt="Language Chart">
                <div class="chart-container" id="monthlyLangChartWrap"><canvas id="monthlyLanguageChart"></canvas></div>
            </div>
            <div class="data-card">
                <h3>📊 Language Data</h3>
                <div class="table-container">
                    <table id="monthlyLangTable">
                        <thead>
                            <tr>
                                <th class="sortable" onclick="sortTable('monthlyLangTable', 0)">Language</th>
                                {''.join([f'<th class="number sortable" onclick="sortTable(' + "'monthlyLangTable', " + str(idx+1) + ')">' + step + '</th>' for idx, step in enumerate(workflow_steps)])}
                                <th class="number sortable" onclick="sortTable('monthlyLangTable', {len(workflow_steps)+1})">Total</th>
                            </tr>
                        </thead>
                        <tbody>
                            {monthly_lang_table_html}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>

        <div class="data-row">
            <div class="data-card">
                <h3>👥 User Productivity</h3>
                <img class="chart-img" id="monthlyUserImg" src="data:image/png;base64,{monthly_user_chart_b64}" alt="User Chart">
                <div class="chart-container" id="monthlyUserChartWrap"><canvas id="monthlyUserChart"></canvas></div>
            </div>
            <div class="data-card">
                <h3>📊 User Data</h3>
                <div class="table-container">
                    <table id="monthlyUserTable">
                        <thead>
                            <tr>
                                <th class="sortable" onclick="sortTable('monthlyUserTable', 0)">Name</th>
                                <th class="sortable" onclick="sortTable('monthlyUserTable', 1)">Language</th>
                                {''.join([f'<th class="number sortable" onclick="sortTable(' + "'monthlyUserTable', " + str(idx+2) + ')">' + step + '</th>' for idx, step in enumerate(workflow_steps)])}
                                <th class="number sortable" onclick="sortTable('monthlyUserTable', {len(workflow_steps)+2})">Total</th>
                            </tr>
                        </thead>
                        <tbody>
                            {monthly_user_table_html}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>

        {self._volunteer_hours_html()}

        {ytd_section_html}

        {self._volunteer_hours_html_ytd()}
    </div>

    <script>
        // User-language mappings
        const monthlyUserLangMap = {json.dumps(monthly_user_lang_map)};
        const ytdUserLangMap = {json.dumps(ytd_user_lang_map)};

        // Initialize Chart.js if available (swap static images for dynamic canvases)
        let monthlyLanguageChart, monthlyUserChart, ytdLanguageChart, ytdUserChart;
        if (typeof Chart !== 'undefined') {{
            // Hide static images, show canvases
            const imgIds = ['monthlyLangImg','monthlyUserImg'{',' + "'ytdLangImg','ytdUserImg'" if not self.weekly else ''}];
            imgIds.forEach(id => {{
                const el = document.getElementById(id);
                if (el) el.style.display = 'none';
            }});
            const chartIds = ['monthlyLangChartWrap','monthlyUserChartWrap'{',' + "'ytdLangChartWrap','ytdUserChartWrap'" if not self.weekly else ''}];
            chartIds.forEach(id => {{
                const el = document.getElementById(id);
                if (el) el.style.display = 'block';
            }});

            const fmtTick = v => v.toLocaleString();

            monthlyLanguageChart = new Chart(document.getElementById('monthlyLanguageChart'), {{
                type: 'bar',
                data: {{ labels: {json.dumps(monthly_lang_labels)}, datasets: {json.dumps(monthly_workflow_datasets)} }},
                options: {{ responsive: true, maintainAspectRatio: false,
                    plugins: {{ title: {{ display: true, text: 'Words by Language & Workflow Step' }}, legend: {{ position: 'top' }} }},
                    scales: {{ x: {{ stacked: true }}, y: {{ stacked: true, beginAtZero: true, ticks: {{ callback: fmtTick }} }} }}
                }}
            }});

            monthlyUserChart = new Chart(document.getElementById('monthlyUserChart'), {{
                type: 'bar',
                data: {{ labels: {json.dumps(monthly_user_labels)}, datasets: [{{ label: 'Total Words', data: {json.dumps(monthly_user_data_points)}, backgroundColor: '#36A2EB' }}] }},
                options: {{ responsive: true, maintainAspectRatio: false,
                    plugins: {{ title: {{ display: true, text: 'Words by User' }}, legend: {{ display: false }} }},
                    scales: {{ y: {{ beginAtZero: true, ticks: {{ callback: fmtTick }} }} }}
                }}
            }});

            {ytd_charts_js}
        }}

        function updateChartFromTable(tableId) {{
            if (typeof Chart === 'undefined') return;
            const table = document.getElementById(tableId);
            const rows = table.querySelectorAll('tbody tr');
            // User charts are labeled "Name (Language)"; rebuild the same identity from the
            // table so a user appears only for the language(s) currently visible. Language
            // charts are labeled by language alone, which matches cell[0] directly.
            const isUserTable = tableId.endsWith('UserTable');
            const visibleLabels = [];
            rows.forEach(row => {{
                if (row.style.display === 'none') return;
                if (isUserTable) {{
                    visibleLabels.push(row.cells[0].textContent.trim() + ' (' + row.cells[1].textContent.trim() + ')');
                }} else {{
                    visibleLabels.push(row.cells[0].textContent.trim());
                }}
            }});

            if (tableId === 'monthlyLangTable') updateBarChart(monthlyLanguageChart, {json.dumps(monthly_lang_labels)}, {json.dumps(monthly_workflow_datasets)}, visibleLabels);
            else if (tableId === 'monthlyUserTable') updateBarChart(monthlyUserChart, {json.dumps(monthly_user_labels)}, [{{ label: 'Total Words', data: {json.dumps(monthly_user_data_points)}, backgroundColor: '#36A2EB' }}], visibleLabels);
            else if (tableId === 'ytdLangTable') updateLineChart(ytdLanguageChart, {json.dumps(ytd_lang_datasets)}, visibleLabels);
            else if (tableId === 'ytdUserTable') updateLineChart(ytdUserChart, {json.dumps(ytd_user_datasets)}, visibleLabels);
        }}

        function updateBarChart(chart, allLabels, allDatasets, visibleLabels) {{
            if (!chart) return;
            const idx = []; allLabels.forEach((l, i) => {{ if (visibleLabels.includes(l)) idx.push(i); }});
            chart.data.labels = idx.map(i => allLabels[i]);
            chart.data.datasets = allDatasets.map(ds => ({{ ...ds, data: idx.map(i => ds.data[i]) }}));
            chart.update();
        }}

        function updateLineChart(chart, allDatasets, visibleLabels) {{
            if (!chart) return;
            chart.data.datasets = allDatasets.filter(ds => visibleLabels.includes(ds.label));
            chart.update();
        }}

        function filterCheckboxes(containerId, searchValue) {{
            const container = document.getElementById(containerId);
            const items = container.querySelectorAll('.checkbox-item');
            const search = searchValue.toUpperCase();

            items.forEach(item => {{
                const label = item.querySelector('label').textContent.toUpperCase();
                item.style.display = label.indexOf(search) > -1 ? '' : 'none';
            }});
        }}

        function applyFilters(section) {{
            // Get checked language checkboxes
            const langCheckboxes = document.querySelectorAll(`#${{section}}LangCheckboxes input[type="checkbox"]:checked`);
            const selectedLangs = Array.from(langCheckboxes).map(cb => cb.value.toUpperCase());

            // Get checked user checkboxes
            const userCheckboxes = document.querySelectorAll(`#${{section}}UserCheckboxes input[type="checkbox"]:checked`);
            const selectedUsers = Array.from(userCheckboxes).map(cb => cb.value.toUpperCase());

            // Update user checkboxes based on selected languages
            updateUserCheckboxes(section, selectedLangs);

            // Filter language table
            const langTable = document.getElementById(`${{section}}LangTable`);
            const langRows = langTable.querySelectorAll('tbody tr');
            langRows.forEach(row => {{
                const langName = row.cells[0].textContent.trim().toUpperCase();
                const show = selectedLangs.length === 0 || selectedLangs.includes(langName);
                row.style.display = show ? '' : 'none';
            }});

            // Filter user table
            const userTable = document.getElementById(`${{section}}UserTable`);
            const userRows = userTable.querySelectorAll('tbody tr');
            userRows.forEach(row => {{
                const userName = row.cells[0].textContent.trim().toUpperCase();
                const userLang = row.cells[1].textContent.trim().toUpperCase();
                let show = true;

                // Apply user checkbox filter
                if (selectedUsers.length > 0 && !selectedUsers.includes(userName)) {{
                    show = false;
                }}

                // Apply language filter to user table
                if (selectedLangs.length > 0 && !selectedLangs.includes(userLang)) {{
                    show = false;
                }}

                row.style.display = show ? '' : 'none';
            }});

            // Update charts
            updateChartFromTable(`${{section}}LangTable`);
            updateChartFromTable(`${{section}}UserTable`);
        }}

        function updateUserCheckboxes(section, selectedLangs) {{
            const userContainer = document.getElementById(`${{section}}UserCheckboxes`);
            const userItems = userContainer.querySelectorAll('.checkbox-item');
            const userLangMap = section === 'monthly' ? monthlyUserLangMap : ytdUserLangMap;
            const searchInput = document.getElementById(`${{section}}UserSearch`);
            const searchVal = searchInput ? searchInput.value.toUpperCase() : '';

            // Build set of users available for selected languages
            const availableUsers = new Set();
            if (selectedLangs.length > 0) {{
                selectedLangs.forEach(langUpper => {{
                    Object.keys(userLangMap).forEach(mapLang => {{
                        if (mapLang.toUpperCase() === langUpper) {{
                            userLangMap[mapLang].forEach(user => availableUsers.add(user.toUpperCase()));
                        }}
                    }});
                }});
            }}

            userItems.forEach(item => {{
                const checkbox = item.querySelector('input[type="checkbox"]');
                const userName = checkbox.value.toUpperCase();
                const labelText = item.querySelector('label').textContent.toUpperCase();

                // Must pass both: language filter AND search filter
                const passesLang = selectedLangs.length === 0 || availableUsers.has(userName);
                const passesSearch = searchVal === '' || labelText.indexOf(searchVal) > -1;

                if (passesLang && passesSearch) {{
                    item.style.display = '';
                }} else {{
                    item.style.display = 'none';
                    if (!passesLang) {{
                        checkbox.checked = false;
                    }}
                }}
            }});
        }}

        function sortTable(tableId, columnIndex) {{
            const table = document.getElementById(tableId);
            const tbody = table.querySelector('tbody');
            const rows = Array.from(tbody.querySelectorAll('tr'));
            const th = table.querySelectorAll('th')[columnIndex];

            const currentSort = th.classList.contains('sort-asc') ? 'asc' :
                               th.classList.contains('sort-desc') ? 'desc' : 'none';
            const newSort = currentSort === 'none' ? 'desc' :
                           currentSort === 'desc' ? 'asc' : 'desc';

            table.querySelectorAll('th').forEach(header => {{
                header.classList.remove('sort-asc', 'sort-desc');
            }});

            if (newSort === 'asc') {{
                th.classList.add('sort-asc');
            }} else {{
                th.classList.add('sort-desc');
            }}

            rows.sort((a, b) => {{
                const aCell = a.cells[columnIndex];
                const bCell = b.cells[columnIndex];

                let aVal = (aCell.dataset.sort !== undefined ? aCell.dataset.sort : aCell.textContent).trim();
                let bVal = (bCell.dataset.sort !== undefined ? bCell.dataset.sort : bCell.textContent).trim();

                aVal = aVal.replace(/,/g, '');
                bVal = bVal.replace(/,/g, '');

                const aNum = parseFloat(aVal);
                const bNum = parseFloat(bVal);

                let comparison = 0;
                if (!isNaN(aNum) && !isNaN(bNum)) {{
                    comparison = aNum - bNum;
                }} else {{
                    comparison = aVal.localeCompare(bVal);
                }}

                return newSort === 'asc' ? comparison : -comparison;
            }});

            rows.forEach(row => tbody.appendChild(row));
            updateChartFromTable(tableId);
        }}
    </script>
</body>
</html>"""

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        return output_path

    def _calculate_summary_stats(self, data: Dict) -> Dict:
        """Calculate summary statistics from the data."""
        stats = {
            'total_words': 0,
            'top_languages': '',
            'workflow_summary': ''
        }

        # Calculate total words and get top languages
        language_totals = {}
        for workflow_key, metrics in data['workflow_by_language'].items():
            language = metrics['language']
            words_done = metrics['words_done']

            if language not in language_totals:
                language_totals[language] = 0
            language_totals[language] += words_done
            stats['total_words'] += words_done

        # Get top 3 languages
        top_languages = sorted(language_totals.items(), key=lambda x: x[1], reverse=True)[:3]
        top_langs_text = []
        for i, (lang, words) in enumerate(top_languages, 1):
            top_langs_text.append(f"  {i}. {lang}: {words:,} words")
        stats['top_languages'] = '\n'.join(top_langs_text) if top_langs_text else "  No data available"

        # Calculate workflow breakdown
        workflow_totals = {}
        for workflow_key, metrics in data['workflow_by_language'].items():
            workflow_step = metrics['workflow_step']
            words_done = metrics['words_done']

            if workflow_step not in workflow_totals:
                workflow_totals[workflow_step] = 0
            workflow_totals[workflow_step] += words_done

        workflow_text = []
        for workflow_step, words in sorted(workflow_totals.items()):
            workflow_text.append(f"- {workflow_step}: {words:,} words")
        stats['workflow_summary'] = '\n'.join(workflow_text) if workflow_text else "No data available"

        return stats

    def _send_system_notification(self, title: str, message: str, sound: bool = True):
        """Send a macOS system notification."""
        try:
            script = f'''
            display notification "{message}" with title "{title}"'''
            if sound:
                script += ' sound name "Glass"'

            subprocess.run(
                ['osascript', '-e', script],
                capture_output=True,
                text=True,
                timeout=5
            )
            logger.info(f"System notification sent: {title}")
        except Exception as e:
            logger.warning(f"Could not send system notification: {e}")

    def _send_failure_notification(self, error_message: str, report_path: str = None):
        """Send notification about failure via multiple channels."""
        logger.info("Sending failure notifications")

        # 1. Send macOS system notification
        self._send_system_notification(
            "XTM Report Failed",
            f"Report generation encountered an error. Check logs for details.",
            sound=True
        )

        # 2. Try to send email notification
        try:
            subject = f"ALERT: XTM Monthly Report Failed - {self.report_month}"
            body = f"""ALERT: The automated XTM monthly report generation has failed.

Report Period: {self.report_month_name}
Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Error: {error_message}

Please check the log file at:
{Path.cwd() / 'xtm_report.log'}
"""
            if report_path:
                body += f"\nPartial report may be available at: {report_path}"

            body += "\n\nPlease investigate and re-run manually if needed."

            # Escape for AppleScript
            subject_escaped = subject.replace('\\', '\\\\').replace('"', '\\"')
            body_escaped = body.replace('\\', '\\\\').replace('"', '\\"')

            # Use error_recipients if specified, otherwise fall back to email_recipients
            recipients = self.config.get('error_recipients', self.config.get('email_recipients', []))
            if recipients:
                # Try simple notification email via Mail app
                recipients_script = ""
                for recipient in recipients:
                    recipient_escaped = recipient.replace('\\', '\\\\').replace('"', '\\"')
                    recipients_script += f'make new to recipient at end of to recipients of new_message with properties {{address:"{recipient_escaped}"}}\n'

                applescript = f'''
                tell application "Mail"
                    set new_message to make new outgoing message with properties {{subject:"{subject_escaped}", visible:false}}
                    tell new_message
                        {recipients_script}
                        set the content to "{body_escaped}"
                    end tell
                    send new_message
                end tell
                '''

                subprocess.run(
                    ['osascript', '-e', applescript],
                    capture_output=True,
                    text=True,
                    timeout=30
                )
                logger.info("Failure notification email sent")
        except Exception as e:
            logger.warning(f"Could not send failure email notification: {e}")

    def _ensure_outlook_running(self) -> bool:
        """Ensure Microsoft Outlook is running, launch if needed."""
        try:
            # Check if Outlook is running
            check_script = '''
            tell application "System Events"
                return (name of processes) contains "Microsoft Outlook"
            end tell
            '''

            result = subprocess.run(
                ['osascript', '-e', check_script],
                capture_output=True,
                text=True,
                timeout=10
            )

            is_running = result.stdout.strip() == "true"

            if is_running:
                logger.info("Microsoft Outlook is already running")
                return True

            # Launch Outlook
            logger.info("Microsoft Outlook is not running. Launching...")
            launch_script = '''
            tell application "Microsoft Outlook"
                activate
            end tell
            '''

            result = subprocess.run(
                ['osascript', '-e', launch_script],
                capture_output=True,
                text=True,
                timeout=30
            )

            if result.returncode == 0:
                # Wait for Outlook to fully start (up to 30 seconds)
                logger.info("Waiting for Outlook to start...")
                import time
                for i in range(30):
                    time.sleep(1)
                    check_result = subprocess.run(
                        ['osascript', '-e', check_script],
                        capture_output=True,
                        text=True,
                        timeout=10
                    )
                    if check_result.stdout.strip() == "true":
                        logger.info(f"Outlook started successfully after {i+1} seconds")
                        # Give it a couple more seconds to fully initialize
                        time.sleep(3)
                        return True

                logger.warning("Outlook did not start within 30 seconds")
                return False
            else:
                logger.warning(f"Failed to launch Outlook: {result.stderr}")
                return False

        except Exception as e:
            logger.warning(f"Error checking/launching Outlook: {e}")
            return False

    def send_email_via_outlook(self, html_path: str, excel_path: str, monthly_data: Dict, ytd_data: Dict):
        """Create and open email draft in Outlook (macOS version) with both HTML and Excel attachments."""
        try:
            logger.info("Preparing to send email via Outlook")

            # Ensure Outlook is running (especially important for auto-send)
            if self.auto_send:
                if not self._ensure_outlook_running():
                    logger.error("Could not ensure Outlook is running. Falling back to Apple Mail.")
                    # Fall through to Apple Mail attempt below

            # Calculate summary statistics
            monthly_stats = self._calculate_summary_stats(monthly_data)

            # Prepare email content
            if self.weekly:
                subject = f"XTM Weekly Report - {self.report_week_label}"
                body = f"""Hello,

Please find attached the XTM weekly report for {self.report_week_label} ({self.report_period}).

Weekly Summary:
- Total Words Processed: {monthly_stats['total_words']:,}
- Top Languages:
{monthly_stats['top_languages']}

Two reports are attached:

1. Interactive HTML Report:
   • Bar charts showing translation volume by language and workflow step
   • User productivity bar charts
   • Sortable tables (click any column header to sort)
   • Filterable data (use the search boxes to filter by language or user)
   • Open in your web browser for full interactivity

2. Excel Workbook (.xlsx):
   • Weekly data sheet with workflow breakdown
   • User statistics sheet
   • Built-in bar charts
   • AutoFilter enabled for easy sorting and filtering
   • Open in Excel, Google Sheets, or Numbers

Report Generated: {self.report_date.strftime('%Y-%m-%d %H:%M')}

Best regards
"""
            else:
                ytd_stats = self._calculate_summary_stats(ytd_data)
                subject = f"XTM Monthly Report - {self.report_month}"
                body = f"""Hello,

Please find attached the XTM monthly report for {self.report_month}.

Monthly Summary ({self.report_month_name}):
- Total Words Processed: {monthly_stats['total_words']:,}
- Top Languages:
{monthly_stats['top_languages']}

Year-to-Date Summary ({self.ytd_start_month} to {self.ytd_end_month}):
- Total Words Processed: {ytd_stats['total_words']:,}
- Top Languages:
{ytd_stats['top_languages']}

Two reports are attached:

1. Interactive HTML Report:
   • Bar charts showing translation volume by language and workflow step
   • User productivity bar charts
   • Line charts showing monthly trends for all languages
   • Sortable tables (click any column header to sort)
   • Filterable data (use the search boxes to filter by language or user)
   • Open in your web browser for full interactivity

2. Excel Workbook (.xlsx):
   • Monthly data sheet with workflow breakdown
   • Year-to-Date sheet with monthly trends
   • Built-in bar and line charts
   • AutoFilter enabled for easy sorting and filtering
   • Open in Excel, Google Sheets, or Numbers

Report Generated: {self.report_date.strftime('%Y-%m-%d %H:%M')}

Best regards
"""

            # Escape quotes and backslashes for AppleScript
            subject_escaped = subject.replace('\\', '\\\\').replace('"', '\\"')
            body_escaped = body.replace('\\', '\\\\').replace('"', '\\"')
            html_path_posix = Path(html_path).as_posix()
            excel_path_posix = Path(excel_path).as_posix()

            # Use weekly_recipients for weekly reports, email_recipients for monthly
            if self.weekly:
                recipients = self.config.get('weekly_recipients', self.config['email_recipients'])
            else:
                recipients = self.config['email_recipients']

            # Try Microsoft Outlook first (already ensured it's running if auto_send)
            outlook_result = self._create_outlook_email_mac(subject_escaped, body_escaped, recipients,
                                                           html_path_posix, excel_path_posix)
            if outlook_result:
                return len(recipients)

            # Fall back to Apple Mail
            logger.info("Microsoft Outlook not available, trying Apple Mail")
            apple_result = self._create_apple_mail_email(subject_escaped, body_escaped, recipients,
                                                        html_path_posix, excel_path_posix)
            if apple_result:
                return len(recipients)

            # If both fail, just open the reports
            logger.warning("Could not create email draft. Opening report files.")
            subprocess.run(['open', html_path])
            subprocess.run(['open', excel_path])
            print(f"\n⚠ Email draft creation failed. Reports opened:")
            print(f"  HTML: {html_path}")
            print(f"  Excel: {excel_path}")
            print(f"Recipients: {', '.join(recipients)}")

        except Exception as e:
            logger.error(f"Failed to create email: {e}")
            logger.info(f"HTML report saved to: {html_path}")
            logger.info(f"Excel report saved to: {excel_path}")
            raise

    def _create_outlook_email_mac(self, subject: str, body: str, recipients: List[str],
                                 html_path: str, excel_path: str) -> bool:
        """Create email draft in Microsoft Outlook for Mac using AppleScript with both attachments."""
        try:
            # Build recipient list for AppleScript - simpler syntax
            recipients_script = ""
            for recipient in recipients:
                recipient_escaped = recipient.replace('\\', '\\\\').replace('"', '\\"')
                recipients_script += f'make new to recipient with properties {{email address:{{address:"{recipient_escaped}"}}}}\n'

            # Different AppleScript depending on whether we're sending or just creating draft
            if self.auto_send:
                applescript = f'''
                tell application "Microsoft Outlook"
                    set new_message to make new outgoing message with properties {{subject:"{subject}", content:"{body}"}}
                    tell new_message
                        {recipients_script}
                        make new attachment with properties {{file:POSIX file "{html_path}"}}
                        make new attachment with properties {{file:POSIX file "{excel_path}"}}
                        send
                    end tell
                end tell
                '''
            else:
                applescript = f'''
                tell application "Microsoft Outlook"
                    set new_message to make new outgoing message with properties {{subject:"{subject}", content:"{body}"}}
                    tell new_message
                        {recipients_script}
                        make new attachment with properties {{file:POSIX file "{html_path}"}}
                        make new attachment with properties {{file:POSIX file "{excel_path}"}}
                    end tell
                    open new_message
                    activate
                end tell
                '''

            result = subprocess.run(
                ['osascript', '-e', applescript],
                capture_output=True,
                text=True,
                timeout=30
            )

            if result.returncode == 0:
                if self.auto_send:
                    logger.info("Microsoft Outlook email sent successfully")
                else:
                    logger.info("Microsoft Outlook email draft created successfully")
                return True
            else:
                logger.warning(f"Outlook AppleScript failed: {result.stderr}")
                return False

        except Exception as e:
            logger.warning(f"Failed to create Outlook email on Mac: {e}")
            return False

    def _create_apple_mail_email(self, subject: str, body: str, recipients: List[str],
                                html_path: str, excel_path: str) -> bool:
        """Create email draft in Apple Mail using AppleScript with both attachments."""
        try:
            # Build recipient list for AppleScript
            recipients_script = ""
            for recipient in recipients:
                recipient_escaped = recipient.replace('\\', '\\\\').replace('"', '\\"')
                recipients_script += f'make new to recipient at end of to recipients of new_message with properties {{address:"{recipient_escaped}"}}\n'

            # Different AppleScript depending on whether we're sending or just creating draft
            # Use simple attachment without positioning to avoid HTML file creation
            if self.auto_send:
                applescript = f'''
                tell application "Mail"
                    set new_message to make new outgoing message with properties {{subject:"{subject}", sender:"leo.chang@familysearch.org"}}
                    tell new_message
                        {recipients_script}
                        set the content to "{body}"
                        make new attachment with properties {{file name:POSIX file "{html_path}"}}
                        make new attachment with properties {{file name:POSIX file "{excel_path}"}}
                    end tell
                    send new_message
                end tell
                '''
            else:
                applescript = f'''
                tell application "Mail"
                    set new_message to make new outgoing message with properties {{subject:"{subject}", visible:true, sender:"leo.chang@familysearch.org"}}
                    tell new_message
                        {recipients_script}
                        set the content to "{body}"
                        make new attachment with properties {{file name:POSIX file "{html_path}"}}
                        make new attachment with properties {{file name:POSIX file "{excel_path}"}}
                    end tell
                    activate
                end tell
                '''

            result = subprocess.run(
                ['osascript', '-e', applescript],
                capture_output=True,
                text=True,
                timeout=30
            )

            if result.returncode == 0:
                if self.auto_send:
                    logger.info("Apple Mail email sent successfully")
                else:
                    logger.info("Apple Mail email draft created successfully")
                return True
            else:
                logger.warning(f"Apple Mail AppleScript failed: {result.stderr}")
                return False

        except Exception as e:
            logger.warning(f"Failed to create Apple Mail email: {e}")
            return False

    def _compute_volunteer_hours(self):
        """Fetch + aggregate volunteer active hours for the reporting period.

        Uses the GUI login/logout history (Playwright helper). Best-effort:
        stores an 'unavailable' summary on any failure so the report still
        generates. Respects EXCLUDED_USERS.
        """
        try:
            import volunteer_hours as vh
        except Exception as e:
            logger.warning("volunteer_hours module unavailable: %s", e)
            self._volunteer_hours = {"by_user": {}, "unavailable": True,
                                     "total_hours": 0.0, "total_sessions": 0,
                                     "volunteer_count": 0}
            return

        if self.weekly:
            # Weekly reports have no YTD; fetch just the week.
            period_start = self.report_start_date.date()
            period_end = self.report_end_date.date()
            logger.info("Fetching volunteer login-history hours for %s..%s",
                        period_start, period_end)
            self._volunteer_hours = vh.get_volunteer_hours(
                period_start, period_end, self.EXCLUDED_USERS, refresh=True)
            self._volunteer_hours_ytd = None
        else:
            year, month = (int(x) for x in self.report_month.split('-'))
            month_start = datetime(year, month, 1).date()
            # last day of the report month = day before the first of next month
            next_month = datetime(year + (month == 12), (month % 12) + 1, 1).date()
            month_end = next_month - timedelta(days=1)
            ytd_start = datetime(year, 1, 1).date()
            ytd_end = month_end

            # Fetch the whole YTD range once; derive the current month from the
            # same fetched data (no second browser login/fetch).
            logger.info("Fetching volunteer login-history hours (YTD) for %s..%s",
                        ytd_start, ytd_end)
            self._volunteer_hours_ytd = vh.get_volunteer_hours(
                ytd_start, ytd_end, self.EXCLUDED_USERS, refresh=True)
            src = self._volunteer_hours_ytd.get("source_file")
            if src and not self._volunteer_hours_ytd.get("unavailable"):
                self._volunteer_hours = vh.aggregate_from_file(
                    src, month_start, month_end, self.EXCLUDED_USERS)
                # Per-month breakdown for the YTD trend chart (Jan..report month).
                months = [f"{year:04d}-{m:02d}" for m in range(1, month + 1)]
                self._volunteer_hours_ytd_breakdown = vh.aggregate_monthly_breakdown(
                    src, months, self.EXCLUDED_USERS)
            else:
                # Fallback: fetch the month on its own if the YTD fetch failed.
                self._volunteer_hours = vh.get_volunteer_hours(
                    month_start, month_end, self.EXCLUDED_USERS, refresh=True)

        for label, summary in (("month", self._volunteer_hours),
                               ("YTD", self._volunteer_hours_ytd)):
            if not summary:
                continue
            if summary.get("unavailable"):
                logger.warning("Volunteer hours (%s) unavailable", label)
            else:
                logger.info("Volunteer hours (%s): %.1fh across %d volunteers (%d sessions)",
                            label, summary.get("total_hours", 0),
                            summary.get("volunteer_count", 0), summary.get("total_sessions", 0))

    def _volunteer_name_map(self) -> Dict[str, str]:
        """Map lowercased login/username -> display name, from the XTM roster.
        Falls back to an empty map (callers default to the login) on failure."""
        try:
            roster = self.get_volunteers()
            return {info.get('username', '').lower(): info.get('user') or info.get('username', '')
                    for info in roster.values() if info.get('username')}
        except Exception as e:
            logger.warning("Could not build volunteer name map: %s", e)
            return {}

    def _volunteer_language_map(self) -> Dict[str, str]:
        """Map lowercased login/username -> comma-joined assigned target
        languages, from the XTM roster. Empty map on failure."""
        try:
            roster = self.get_volunteers()
            return {info.get('username', '').lower(): ', '.join(info.get('languages') or [])
                    for info in roster.values() if info.get('username')}
        except Exception as e:
            logger.warning("Could not build volunteer language map: %s", e)
            return {}

    def _volunteer_hours_html(self) -> str:
        """Render the current-period Volunteer Hours HTML section."""
        period = (self.report_period if self.weekly
                  else (self.report_month_name or self.report_month))
        return self._render_volunteer_hours_html(
            self._volunteer_hours, f"Volunteer Hours in XTM - {period}", "volunteerHoursTable")

    def _volunteer_hours_html_ytd(self) -> str:
        """Render the YTD Volunteer Hours HTML section: per-month breakdown table
        plus a line chart of the top-10 volunteers' monthly trends (mirrors the
        other year-to-date sections)."""
        if self.weekly:
            return ""
        heading = f"⏱️ Volunteer Hours in XTM - Year-to-Date ({self.ytd_start_month} to {self.ytd_end_month})"
        bd = self._volunteer_hours_ytd_breakdown
        if not bd or not bd.get("by_user"):
            summ = self._volunteer_hours_ytd
            if summ and summ.get("unavailable"):
                return f"""
        <div class="section-divider"><h2>{heading}</h2></div>
        <div class="data-card"><p style="color:#a00;">Volunteer login-history hours were
        unavailable for the year-to-date period. All other data is unaffected.</p></div>"""
            return ""

        import volunteer_hours as _vh
        months = bd["months"]
        by_user = bd["by_user"]
        name_by_login = self._volunteer_name_map()
        lang_by_login = self._volunteer_language_map()
        sorted_users = sorted(by_user.items(), key=lambda x: -x[1]["active_seconds"])

        # Static line chart (hours per month), same helper as the other YTD charts.
        chart_ds = [{'label': f"{name_by_login.get(login.lower(), login)} ({lang_by_login.get(login.lower(), '')})",
                     'data': [round(ud['months'].get(m, 0) / 3600.0, 2) for m in months]}
                    for login, ud in sorted_users[:10]]
        chart_b64 = self._generate_line_chart_base64(months, chart_ds, 'Volunteer Hours Trends (Top 10)')

        month_headers = "".join(
            f"<th class=\"number sortable\" onclick=\"sortTable('volunteerHoursYtdTable', {i + 2})\">{m}</th>"
            for i, m in enumerate(months))
        total_idx = len(months) + 2
        rows_html = ""
        for login, ud in sorted_users:
            cells = "".join(
                f"<td class=\"number\" data-sort=\"{ud['months'].get(m, 0)}\">{_vh.format_hms(ud['months'].get(m, 0))}</td>"
                for m in months)
            rows_html += (
                f"<tr><td>{name_by_login.get(login.lower(), login)}</td>"
                f"<td>{lang_by_login.get(login.lower(), '')}</td>"
                f"{cells}"
                f"<td class=\"number\" data-sort=\"{ud['active_seconds']}\">{_vh.format_hms(ud['active_seconds'])}</td></tr>")

        total_hms = _vh.format_hms(self._volunteer_hours_ytd.get('total_seconds', 0)
                                   if self._volunteer_hours_ytd else 0)
        vol_count = self._volunteer_hours_ytd.get('volunteer_count', len(by_user)) if self._volunteer_hours_ytd else len(by_user)
        return f"""
        <div class="section-divider">
            <h2>{heading}</h2>
        </div>
        <div class="data-row">
            <div class="data-card">
                <h3>📈 Hours Trend</h3>
                <img class="chart-img" src="data:image/png;base64,{chart_b64}" alt="YTD Volunteer Hours Chart">
                <p style="font-size:0.85em;color:#666;">Total YTD active time: {total_hms} (h:mm:ss)
                across {vol_count} volunteers. Active time = login to last action per session.</p>
            </div>
            <div class="data-card">
                <h3>📊 Monthly Hours by Volunteer</h3>
                <div class="table-container">
                    <table id="volunteerHoursYtdTable">
                        <thead>
                            <tr>
                                <th class="sortable" onclick="sortTable('volunteerHoursYtdTable', 0)">Volunteer</th>
                                <th class="sortable" onclick="sortTable('volunteerHoursYtdTable', 1)">Language</th>
                                {month_headers}
                                <th class="number sortable" onclick="sortTable('volunteerHoursYtdTable', {total_idx})">Total (h:mm:ss)</th>
                            </tr>
                        </thead>
                        <tbody>
                            {rows_html}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>"""

    def _render_volunteer_hours_html(self, vh: Dict, heading: str, table_id: str) -> str:
        """Render a Volunteer Hours HTML section from a summary dict."""
        if not vh:
            return ""
        if vh.get("unavailable") or not vh.get("by_user"):
            return f"""
        <div class="section-divider">
            <h2>⏱️ {heading}</h2>
        </div>
        <div class="data-card">
            <p style="color:#a00;">Volunteer login-history hours were unavailable for this
            period (the interactive XTM login fetch did not complete). All other data is unaffected.</p>
        </div>"""

        import volunteer_hours as _vh
        name_by_login = self._volunteer_name_map()
        lang_by_login = self._volunteer_language_map()
        rows = sorted(vh["by_user"].items(), key=lambda x: -x[1]["active_seconds"])
        body = "".join(
            f"<tr><td>{name_by_login.get(login.lower(), login)}</td>"
            f"<td>{lang_by_login.get(login.lower(), '')}</td>"
            f"<td class=\"number\" data-sort=\"{d['active_seconds']}\">{_vh.format_hms(d['active_seconds'])}</td>"
            f"<td class=\"number\">{d['sessions']}</td></tr>"
            for login, d in rows
        )
        total_hms = _vh.format_hms(vh.get("total_seconds", 0))
        return f"""
        <div class="section-divider">
            <h2>⏱️ {heading}</h2>
        </div>
        <div class="data-row">
            <div class="data-card">
                <h3>📈 Summary</h3>
                <p><strong>Total active time:</strong> {total_hms} (h:mm:ss)<br>
                <strong>Volunteers active:</strong> {vh['volunteer_count']}<br>
                <strong>Total login sessions:</strong> {vh['total_sessions']:,}</p>
                <p style="font-size:0.85em;color:#666;">Active time = login to the last recorded
                action in each session (excludes idle time before logout). Source: XTM
                login/logout history (minute precision, so seconds show as 00).</p>
            </div>
            <div class="data-card">
                <h3>📊 Time by Volunteer</h3>
                <div class="table-container">
                    <table id="{table_id}">
                        <thead>
                            <tr>
                                <th class="sortable" onclick="sortTable('{table_id}', 0)">Volunteer</th>
                                <th class="sortable" onclick="sortTable('{table_id}', 1)">Language</th>
                                <th class="number sortable" onclick="sortTable('{table_id}', 2)">Active Time (h:mm:ss)</th>
                                <th class="number sortable" onclick="sortTable('{table_id}', 3)">Sessions</th>
                            </tr>
                        </thead>
                        <tbody>
                            {body}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>"""

    def _create_volunteer_hours_sheet(self, wb):
        """Add the current-period Volunteer Hours sheet, and (monthly) a YTD one."""
        title = ("Weekly" if self.weekly else "Monthly") + " Volunteer Hours in XTM"
        self._write_volunteer_hours_sheet(wb, self._volunteer_hours, "Volunteer Hours", title)
        if not self.weekly:
            self._write_volunteer_hours_ytd_sheet(wb)

    def _write_volunteer_hours_ytd_sheet(self, wb):
        """YTD Volunteer Hours sheet: per-month breakdown + line chart of the
        top-10 volunteers' monthly trends (mirrors the 'User Stats - YTD' sheet)."""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.series import SeriesLabel
        from openpyxl.utils import get_column_letter

        bd = self._volunteer_hours_ytd_breakdown
        if not bd or not bd.get("by_user"):
            return
        months = bd["months"]
        by_user = bd["by_user"]
        name_by_login = self._volunteer_name_map()
        lang_by_login = self._volunteer_language_map()

        ws = wb.create_sheet("Volunteer Hours YTD")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        headers = ["Volunteer", "Language"] + months + ["Total"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        sorted_users = sorted(by_user.items(),
                              key=lambda x: x[1]["active_seconds"], reverse=True)
        total_col = len(headers)
        for row_idx, (login, ud) in enumerate(sorted_users, 2):
            ws.cell(row=row_idx, column=1, value=name_by_login.get(login.lower(), login))
            ws.cell(row=row_idx, column=2, value=lang_by_login.get(login.lower(), ""))
            for col_idx, month in enumerate(months, 3):
                secs = ud["months"].get(month, 0)
                # Real Excel duration (fraction of a day) shown as [h]:mm:ss.
                c = ws.cell(row=row_idx, column=col_idx, value=secs / 86400.0)
                c.number_format = "[h]:mm:ss"
                c.alignment = Alignment(horizontal="right")
            tcell = ws.cell(row=row_idx, column=total_col, value=ud["active_seconds"] / 86400.0)
            tcell.number_format = "[h]:mm:ss"
            tcell.font = Font(bold=True)
            tcell.fill = PatternFill(start_color="F0F7FF", end_color="F0F7FF", fill_type="solid")
            tcell.alignment = Alignment(horizontal="right")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 24
        for col_idx in range(3, total_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
        ws.auto_filter.ref = ws.dimensions

        # Line chart: top 10 volunteers' monthly active-time trends.
        num_users = min(10, len(sorted_users))
        if num_users > 0:
            chart = LineChart()
            chart.title = "Volunteer Hours Trends (Top 10)"
            chart.y_axis.title = "Active time (h:mm:ss)"
            chart.x_axis.title = "Month"
            for row_idx in range(2, num_users + 2):
                data = Reference(ws, min_col=3, max_col=len(months) + 2,
                                 min_row=row_idx, max_row=row_idx)
                chart.add_data(data, titles_from_data=False)
            cats = Reference(ws, min_col=3, max_col=len(months) + 2, min_row=1, max_row=1)
            chart.set_categories(cats)
            for idx, (login, ud) in enumerate(sorted_users[:num_users]):
                sl = SeriesLabel()
                sl.strRef = None
                sl.v = f"{name_by_login.get(login.lower(), login)} ({lang_by_login.get(login.lower(), '')})"
                chart.series[idx].tx = sl
            chart.height = 15
            chart.width = 25
            ws.add_chart(chart, f"A{len(sorted_users) + 3}")

    def _write_volunteer_hours_sheet(self, wb, vh, sheet_name, title):
        """Write one Volunteer Hours sheet from a summary dict."""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference

        if not vh or vh.get("unavailable") or not vh.get("by_user"):
            return

        import volunteer_hours as _vh
        name_by_login = self._volunteer_name_map()
        lang_by_login = self._volunteer_language_map()

        ws = wb.create_sheet(sheet_name)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=13)
        ws["A2"] = (f"Total active time: {_vh.format_hms(vh.get('total_seconds', 0))} (h:mm:ss)   |   "
                    f"Volunteers: {vh['volunteer_count']}   |   "
                    f"Sessions: {vh['total_sessions']:,}")
        ws["A3"] = ("Active time = login to last action in session (excludes idle time before "
                    "logout). Source: XTM login/logout history (minute precision).")
        ws["A3"].font = Font(italic=True, size=9, color="666666")

        header_row = 5
        # Column E holds numeric hours purely to drive the chart (hidden).
        headers = ["Volunteer", "Language", "Active Time (h:mm:ss)", "Sessions", "Hours (chart)"]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")

        rows = sorted(vh["by_user"].items(), key=lambda x: -x[1]["active_seconds"])
        r = header_row + 1
        for login, d in rows:
            ws.cell(row=r, column=1, value=name_by_login.get(login.lower(), login))
            ws.cell(row=r, column=2, value=lang_by_login.get(login.lower(), ""))
            # Real Excel duration: value in days, displayed as [h]:mm:ss.
            tcell = ws.cell(row=r, column=3, value=d["active_seconds"] / 86400.0)
            tcell.number_format = "[h]:mm:ss"
            ws.cell(row=r, column=4, value=d["sessions"])
            ws.cell(row=r, column=5, value=round(d["active_seconds"] / 3600.0, 2))
            r += 1

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].hidden = True
        ws.auto_filter.ref = f"A{header_row}:D{r - 1}"

        # Bar chart of top 20 volunteers by active hours (numeric helper column E)
        n = min(20, len(rows))
        if n >= 1:
            chart = BarChart()
            chart.title = "Top Volunteers by Active Hours"
            chart.type = "bar"
            chart.y_axis.title = "Hours"
            data = Reference(ws, min_col=5, min_row=header_row, max_row=header_row + n)
            cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + n)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 12
            chart.width = 22
            ws.add_chart(chart, "G5")

    def generate_report(self):
        """Main method to generate and distribute the HTML/Excel report."""
        html_path = None
        try:
            if self.weekly:
                logger.info(f"Starting XTM weekly report generation for {self.report_week_label}")
            else:
                logger.info("Starting XTM monthly report generation")

            # Run health checks
            self._run_health_checks()

            if self.weekly:
                # Query API for weekly data
                try:
                    weekly_data = self.aggregate_weekly_data(self.report_start_date, self.report_end_date)
                except Exception as e:
                    logger.error(f"Failed to aggregate weekly data: {e}", exc_info=True)
                    weekly_data = {
                        'project_stats': {'total': 0, 'completed': 0, 'in_progress': 0, 'pending': 0},
                        'workflow_by_language': {},
                        'user_statistics': {},
                        'projects': []
                    }

                monthly_data = weekly_data
                # No YTD for weekly reports
                ytd_monthly_breakdown = {'months': [], 'languages': {}, 'users': {}}
            else:
                # Query API for current month data
                try:
                    monthly_data = self.aggregate_monthly_data(self.report_month, self.report_month)
                    # Cache for future runs
                    self._save_month_cache(self.report_month, monthly_data)
                except Exception as e:
                    logger.error(f"Failed to aggregate monthly data: {e}", exc_info=True)
                    monthly_data = {
                        'project_stats': {'total': 0, 'completed': 0, 'in_progress': 0, 'pending': 0},
                        'workflow_by_language': {},
                        'user_statistics': {},
                        'projects': []
                    }

                # Aggregate YTD data (reuses monthly_data for current month, cache for past months)
                try:
                    ytd_monthly_breakdown = self.aggregate_ytd_data(
                        self.ytd_start_month, self.ytd_end_month,
                        current_month_data=monthly_data
                    )
                except Exception as e:
                    logger.error(f"Failed to aggregate YTD data: {e}", exc_info=True)
                    ytd_monthly_breakdown = {'months': [], 'languages': {}, 'users': {}}

            # Add every volunteer in the XTM roster (minus excluded users) to the
            # user report, including those with no work this period (zero rows,
            # one per assigned language). Done after caching/YTD so the month
            # caches keep only real work. Charts stay limited to non-zero users.
            try:
                self._inject_zero_volunteers(monthly_data.get('user_statistics', {}), 'workflow_steps')
                if not self.weekly:
                    self._inject_zero_volunteers(ytd_monthly_breakdown.get('users', {}), 'months')
                    # Keep the YTD language table in sync with the user report:
                    # include every language volunteers represent, even idle ones.
                    ytd_langs = ytd_monthly_breakdown.setdefault('languages', {})
                    for lang in self._report_language_set(ytd_monthly_breakdown):
                        ytd_langs.setdefault(lang, {})
            except Exception as e:
                logger.warning(f"Failed to add full volunteer roster to user report: {e}")

            # Volunteer time-in-XTM (from GUI login/logout history via Playwright).
            # Best-effort: never blocks report generation.
            self._compute_volunteer_hours()

            if not (monthly_data.get('workflow_by_language') or ytd_monthly_breakdown.get('languages')):
                logger.warning("No data available for reporting period")

            # Ensure output directory exists
            output_dir = Path(self.config['onedrive_path'])
            try:
                output_dir.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                logger.warning(f"Could not create output directory: {e}")

            # Create HTML and Excel reports with appropriate naming
            if self.weekly:
                report_label = f"Weekly_Report_{self.report_week_label.replace('Week of ', '')}"
                html_filename = f"XTM_{report_label}_{self.report_date.strftime('%Y%m%d')}.html"
                excel_filename = f"XTM_{report_label}_{self.report_date.strftime('%Y%m%d')}.xlsx"
            else:
                html_filename = f"XTM_Report_{self.report_month}_{self.report_date.strftime('%Y%m%d')}.html"
                excel_filename = f"XTM_Report_{self.report_month}_{self.report_date.strftime('%Y%m%d')}.xlsx"

            # Save locally first (reliable), then copy to OneDrive
            local_dir = Path(__file__).parent / "output"
            local_dir.mkdir(exist_ok=True)

            local_html_path = str(local_dir / html_filename)
            local_excel_path = str(local_dir / excel_filename)

            self.create_combined_html_report(monthly_data, ytd_monthly_breakdown, local_html_path)
            logger.info(f"Combined HTML report saved to {local_html_path}")

            self.create_excel_report(monthly_data, ytd_monthly_breakdown, local_excel_path)
            logger.info(f"Excel report saved to {local_excel_path}")

            # Copy to OneDrive (best-effort)
            html_path = str(output_dir / html_filename)
            excel_path = str(output_dir / excel_filename)
            import shutil
            try:
                shutil.copy2(local_html_path, html_path)
                shutil.copy2(local_excel_path, excel_path)
                logger.info(f"Reports copied to OneDrive: {output_dir}")
            except Exception as e:
                logger.warning(f"Could not copy to OneDrive (files saved locally): {e}")
                html_path = local_html_path
                excel_path = local_excel_path

            # Construct a minimal ytd_data for the email body stats
            ytd_data = {
                'project_stats': {'total': 0, 'completed': 0, 'in_progress': 0, 'pending': 0},
                'workflow_by_language': {},
                'user_statistics': {},
                'projects': []
            }
            # Populate from YTD breakdown for email summary
            for language, month_words in ytd_monthly_breakdown.get('languages', {}).items():
                total = sum(month_words.values())
                ytd_data['workflow_by_language'][f"translate - {language}"] = {
                    'workflow_step': 'translate', 'language': language,
                    'words_done': total, 'words_to_be_done': 0, 'projects': 0
                }

            # Send via Outlook
            recipient_count = 0
            try:
                recipient_count = self.send_email_via_outlook(html_path, excel_path, monthly_data, ytd_data)
                email_success = True
            except Exception as e:
                logger.error(f"Failed to send email: {e}", exc_info=True)
                email_success = False

            logger.info("Report generation completed successfully")
            print(f"\n✓ Reports generated:")
            print(f"  - HTML: {html_path}")
            print(f"  - Excel: {excel_path}")
            if self.weekly:
                print(f"✓ Period: {self.report_period}")
            else:
                print(f"✓ Period: {self.report_month} (YTD: {self.ytd_start_month} to {self.ytd_end_month})")
            if email_success:
                if self.auto_send:
                    print(f"✓ Email sent automatically to {recipient_count} recipient{'s' if recipient_count != 1 else ''}")
                else:
                    print(f"✓ Email draft created with {recipient_count} recipient{'s' if recipient_count != 1 else ''}")
            else:
                print(f"⚠ Email could not be sent - reports saved locally")

        except Exception as e:
            logger.error(f"Report generation failed: {e}", exc_info=True)
            try:
                self._send_failure_notification(str(e), html_path)
            except:
                pass
            raise


def main():
    """Main entry point."""
    import argparse

    parser = argparse.ArgumentParser(description='Generate XTM report (monthly or weekly)')
    parser.add_argument('--auto-send', action='store_true',
                       help='Automatically send email (default: create draft only)')
    parser.add_argument('--weekly', action='store_true',
                       help='Generate weekly report for previous 7 days (default: monthly report)')
    parser.add_argument('--snapshot', action='store_true',
                       help='Snapshot per-user statistics for all active projects (no report). '
                            'Run regularly so archived projects keep real user names.')
    args = parser.parse_args()

    try:
        generator = XTMReportGenerator(auto_send=args.auto_send, weekly=args.weekly)
        if args.snapshot:
            count = generator.snapshot_active_projects()
            print(f"✓ Snapshotted per-user statistics for {count} active project{'s' if count != 1 else ''}")
            return
        generator.generate_report()
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
